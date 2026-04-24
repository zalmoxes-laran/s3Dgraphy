"""
UnifiedXLSXExporter — inverse of UnifiedXLSXImporter.

Walks an in-memory s3dgraphy Graph and produces a single ``em_data.xlsx``
file with the 5 typed sheets described in
``s3dgraphy/templates/em_data_template.xlsx`` (Units, Epochs, Claims,
Authors, Documents).

Primary use cases:

* **Round-trip testing**. Import graphml or xlsx → export → re-import →
  verify the resolver output and the claim structure are unchanged.
* **Graph-to-xlsx refactoring**. After editing a graph in memory (e.g.
  applying ``compact_propagative_metadata`` or a diagnostics fix),
  persist it back to xlsx so StratiMiner / a human editor can iterate.
* **Migration**. Import a legacy two-file workflow (MappedXLSXImporter
  + QualiaImporter) into memory, export with this writer to produce a
  fresh unified file.

The exporter is deliberately conservative: it emits only what the graph
actually contains. Relations that were created as auto-inferred
``is_overlain_by`` / ``is_cut_by`` (reverse edges) are skipped — only
the primary direction (``overlies``, ``cuts``, …) is written, matching
what the prompt v5.0 recommends.
"""

import os
import re
from typing import Any, Dict, List, Optional, Set, Tuple

from ..graph import Graph
from ..nodes.stratigraphic_node import StratigraphicNode
from ..nodes.epoch_node import EpochNode
from ..nodes.author_node import AuthorNode, AuthorAINode
from ..nodes.document_node import DocumentNode
from ..nodes.extractor_node import ExtractorNode
from ..nodes.combiner_node import CombinerNode
from ..nodes.property_node import PropertyNode


# Only the canonical direction for each pair gets emitted. Reverse edges
# (is_overlain_by, is_cut_by, is_filled_by, is_abutted_by) are recovered
# at import time from the canonical direction, so we skip them here to
# avoid double emission.
_CANONICAL_RELATIONS = frozenset({
    "overlies", "cuts", "fills", "abuts",
    "bonded_to", "equals",
    "is_after", "is_before",
    "has_same_time", "contrasts_with", "changed_from",
})

_REVERSE_RELATIONS = frozenset({
    "is_overlain_by", "is_cut_by", "is_filled_by", "is_abutted_by",
})

_ALL_RELATIONS = _CANONICAL_RELATIONS | _REVERSE_RELATIONS


def _parse_author_description(desc: str) -> Tuple[str, str, str]:
    """Split an AuthorNode.description produced by UnifiedXLSXImporter
    back into ``(display_name, orcid, affiliation)``.

    The importer joins them with ``" | "``. ORCID entries look like
    ``ORCID:0000-…``; anything that isn't an ORCID and isn't the first
    chunk is treated as affiliation.
    """
    if not desc:
        return "", "", ""
    parts = [p.strip() for p in desc.split(" | ") if p.strip()]
    display = parts[0] if parts else ""
    orcid = ""
    affil = ""
    for p in parts[1:]:
        if p.upper().startswith("ORCID:"):
            orcid = p.split(":", 1)[1].strip()
        else:
            affil = p
    return display, orcid, affil


def _parse_document_description(desc: str) -> Tuple[str, str, str]:
    """Split a DocumentNode.description produced by UnifiedXLSXImporter
    into ``(filename, title, year)``. Join separator is ``" | "``.
    """
    if not desc:
        return "", "", ""
    parts = [p.strip() for p in desc.split(" | ") if p.strip()]
    filename = parts[0] if len(parts) > 0 else ""
    title = parts[1] if len(parts) > 1 else ""
    year = parts[2] if len(parts) > 2 else ""
    return filename, title, year


class UnifiedXLSXExporter:
    """Export an in-memory Graph to an ``em_data.xlsx`` workbook.

    Usage::

        exporter = UnifiedXLSXExporter(graph)
        exporter.write('em_data.xlsx')

    The ``write`` method raises ``IOError`` on filesystem problems and
    returns a report dict with how many rows were written to each sheet.
    """

    def __init__(self, graph: Graph):
        self.graph = graph
        # id_by_node_uuid maps a node.node_id to the short code used in
        # the output sheet. For units/epochs/authors/documents the
        # short code IS node.name; for ExtractorNode it's also
        # node.name (D.XX.YY). We keep this indirection to be robust
        # against missing names.
        self._auth_id_by_uuid: Dict[str, str] = {}
        self._doc_id_by_uuid: Dict[str, str] = {}
        self._epoch_id_by_uuid: Dict[str, str] = {}
        self._unit_id_by_uuid: Dict[str, str] = {}

    # ------------------------------------------------------------------
    # Public entry point
    # ------------------------------------------------------------------

    def write(self, path: str) -> Dict[str, int]:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = Workbook()
        wb.remove(wb.active)

        # Headers + order MUST match the template.
        sheets_spec = [
            ("Units",     ["ID", "TYPE", "NAME"]),
            ("Epochs",    ["ID", "NAME", "START", "END", "COLOR"]),
            ("Claims",    ["TARGET_ID", "TARGET2_ID", "PROPERTY_TYPE",
                           "VALUE", "UNITS", "COMBINER_REASONING",
                           "EXTRACTOR_1", "DOCUMENT_1", "AUTHOR_1", "AUTHOR_KIND_1",
                           "EXTRACTOR_2", "DOCUMENT_2", "AUTHOR_2", "AUTHOR_KIND_2"]),
            ("Authors",   ["ID", "KIND", "DISPLAY_NAME", "ORCID", "AFFILIATION"]),
            ("Documents", ["ID", "FILENAME", "TITLE", "YEAR", "AUTHOR_IDS"]),
        ]
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", start_color="1F4E78")
        header_align = Alignment(horizontal="center", vertical="center")

        def add_sheet(name, headers, rows):
            ws = wb.create_sheet(name)
            ws.append(headers)
            for col in range(1, len(headers) + 1):
                c = ws.cell(row=1, column=col)
                c.font = header_font
                c.fill = header_fill
                c.alignment = header_align
            ws.freeze_panes = "A2"
            for row in rows:
                ws.append(row)

        units_rows = self._build_units_rows()
        epochs_rows = self._build_epochs_rows()
        authors_rows = self._build_authors_rows()
        documents_rows = self._build_documents_rows()
        claims_rows = self._build_claims_rows()

        for (name, headers), rows in zip(
                sheets_spec,
                (units_rows, epochs_rows, claims_rows, authors_rows, documents_rows)):
            add_sheet(name, headers, rows)

        # Ensure parent dir exists
        parent = os.path.dirname(os.path.abspath(path))
        if parent:
            os.makedirs(parent, exist_ok=True)
        wb.save(path)

        return {
            "units": len(units_rows),
            "epochs": len(epochs_rows),
            "claims": len(claims_rows),
            "authors": len(authors_rows),
            "documents": len(documents_rows),
        }

    # ------------------------------------------------------------------
    # Sheet builders
    # ------------------------------------------------------------------

    def _build_units_rows(self) -> List[List[Any]]:
        rows = []
        seen_ids: Set[str] = set()
        for node in self.graph.nodes:
            # StratigraphicNode covers US / USVn / USVs / VSF / SF / USD / ...
            if not isinstance(node, StratigraphicNode):
                continue
            base_uid = node.name or node.node_id
            uid = base_uid
            # When two stratigraphic nodes share the same short name
            # (a data-quality issue in the source GraphML, or the
            # generic ``continuity_node`` label on BR nodes),
            # disambiguate with a short uuid suffix so the Units sheet
            # stays unique and each row round-trips to a distinct
            # in-memory node.
            if uid in seen_ids:
                uid = f"{base_uid}__{node.node_id[:6]}"
                self.graph.warnings.append(
                    f"[xlsx export] Duplicate unit name '{base_uid}'; "
                    f"disambiguated as '{uid}'."
                )
            seen_ids.add(uid)
            self._unit_id_by_uuid[node.node_id] = uid
            type_str = getattr(node, "node_type", None) or type(node).__name__
            name_display = getattr(node, "description", "") or ""
            rows.append([uid, type_str, name_display])
        # Deterministic output
        rows.sort(key=lambda r: r[0])
        return rows

    def _build_epochs_rows(self) -> List[List[Any]]:
        rows = []
        for node in self.graph.nodes:
            if not isinstance(node, EpochNode):
                continue
            eid = node.name or node.node_id
            self._epoch_id_by_uuid[node.node_id] = eid
            start = getattr(node, "start_time", None)
            end = getattr(node, "end_time", None)
            color = ""
            if hasattr(node, "color") and node.color:
                color = node.color
            elif hasattr(node, "attributes") and node.attributes.get("fill_color"):
                color = node.attributes["fill_color"]
            rows.append([eid, eid, start, end, color])
        rows.sort(key=lambda r: r[0])
        return rows

    def _build_authors_rows(self) -> List[List[Any]]:
        rows = []
        for node in self.graph.nodes:
            if not isinstance(node, AuthorNode):
                continue
            aid = node.name or node.node_id
            self._auth_id_by_uuid[node.node_id] = aid
            kind = "extractor" if isinstance(node, AuthorAINode) else "author"
            display, orcid, affil = _parse_author_description(
                getattr(node, "description", "") or "")
            # Fall back to the short code when no richer description exists
            if not display:
                display = aid
            rows.append([aid, kind, display, orcid, affil])
        rows.sort(key=lambda r: r[0])
        return rows

    def _build_documents_rows(self) -> List[List[Any]]:
        rows = []
        for node in self.graph.nodes:
            if not isinstance(node, DocumentNode):
                continue
            did = node.name or node.node_id
            self._doc_id_by_uuid[node.node_id] = did
            filename, title, year = _parse_document_description(
                getattr(node, "description", "") or "")
            # Collect has_author edges from the document (doc-level authors)
            author_ids = []
            for edge in self.graph.edges:
                if edge.edge_source == node.node_id and edge.edge_type == "has_author":
                    aid = self._auth_id_by_uuid.get(edge.edge_target)
                    if aid:
                        author_ids.append(aid)
            rows.append([did, filename, title, year, ", ".join(author_ids)])
        rows.sort(key=lambda r: r[0])
        return rows

    # ------------------------------------------------------------------
    # Claims builder — the interesting part
    # ------------------------------------------------------------------

    def _build_claims_rows(self) -> List[List[Any]]:
        rows: List[List[Any]] = []

        # 1. belongs_to_epoch rows (one per has_first_epoch edge)
        for edge in self.graph.edges:
            if edge.edge_type != "has_first_epoch":
                continue
            src = self._unit_id_by_uuid.get(edge.edge_source)
            tgt = self._epoch_id_by_uuid.get(edge.edge_target)
            if not src or not tgt:
                continue
            rows.append(self._row(src, tgt, "belongs_to_epoch"))

        # 2. Stratigraphic relations (canonical direction only).
        for edge in self.graph.edges:
            if edge.edge_type not in _CANONICAL_RELATIONS:
                continue
            src = (self._unit_id_by_uuid.get(edge.edge_source)
                   or self._epoch_id_by_uuid.get(edge.edge_source))
            tgt = (self._unit_id_by_uuid.get(edge.edge_target)
                   or self._epoch_id_by_uuid.get(edge.edge_target))
            if not src or not tgt:
                continue
            row = self._row(src, tgt, edge.edge_type)
            self._fill_attribution_from_edge(row, edge)
            rows.append(row)

        # 3. Qualia claims — one row per has_property edge.
        # For each PN we rebuild the attribution chain.
        #
        # GraphML quirk: the legacy importer sets every PropertyNode's
        # ``property_type`` to the generic sentinel ``"string"`` and
        # stores the real qualia type in ``pn.name``; numeric values
        # often end up in ``pn.description`` (value stays empty).
        # The writer normalizes both.
        for edge in self.graph.edges:
            if edge.edge_type != "has_property":
                continue
            pn = self.graph.find_node_by_id(edge.edge_target)
            if not isinstance(pn, PropertyNode):
                continue
            host_id = (self._unit_id_by_uuid.get(edge.edge_source)
                       or self._epoch_id_by_uuid.get(edge.edge_source))
            if not host_id:
                continue

            # Prefer pn.name whenever property_type is missing or the
            # generic "string" sentinel.
            prop_type = pn.property_type
            if not prop_type or prop_type == "string":
                prop_type = pn.name or "definition"

            # Pull value from value, falling back to description for
            # legacy GraphML storage.
            value = pn.value or pn.description or ""

            units = ""
            if hasattr(pn, "attributes") and pn.attributes:
                units = pn.attributes.get("units") or ""

            # Skip completely empty PropertyNodes (no name, no value, no
            # attribution) — they're parser noise and would break the
            # duplicate-triple check at validation time.
            has_attribution = bool(
                self._first_has_author(pn.node_id)
                or any(e.edge_source == pn.node_id
                       and e.edge_type == "has_data_provenance"
                       for e in self.graph.edges)
            )
            if not value and not has_attribution and prop_type in ("string", "definition"):
                continue

            row = self._row(host_id, "", prop_type, value=value, units=units)
            self._fill_attribution_from_property_chain(row, pn)
            rows.append(row)

        # Deterministic order: group by target_id, then property, then
        # target2 (relations with a specific second endpoint come later).
        rows.sort(key=lambda r: (r[0], r[2], r[1]))
        return rows

    # ------------------------------------------------------------------
    # Row helpers
    # ------------------------------------------------------------------

    def _row(self, target_id, target2_id, property_type,
             value="", units=""):
        return [target_id, target2_id, property_type, value, units,
                "",   # COMBINER_REASONING
                "", "", "", "",   # EXTRACTOR_1/DOCUMENT_1/AUTHOR_1/AUTHOR_KIND_1
                "", "", "", ""]   # EXTRACTOR_2/DOCUMENT_2/AUTHOR_2/AUTHOR_KIND_2

    # Column indices inside a Claims row (0-based) — kept in sync with
    # the header list in write().
    _COL_COMBINER = 5
    _COL_EXT_1 = 6
    _COL_DOC_1 = 7
    _COL_AUTH_1 = 8
    _COL_KIND_1 = 9
    _COL_EXT_2 = 10
    _COL_DOC_2 = 11
    _COL_AUTH_2 = 12
    _COL_KIND_2 = 13

    def _fill_attribution_from_edge(self, row, edge) -> None:
        """Relational claims: attribution was stored on ``edge.attributes``
        by the importer. Copy it back into the row.
        """
        attrs = getattr(edge, "attributes", {}) or {}
        for suffix, auth_col, kind_col, doc_col in (
            ("_1", self._COL_AUTH_1, self._COL_KIND_1, self._COL_DOC_1),
            ("_2", self._COL_AUTH_2, self._COL_KIND_2, self._COL_DOC_2),
        ):
            aid = attrs.get(f"authored_by{suffix}", "") or ""
            if not aid:
                continue
            row[auth_col] = aid
            row[kind_col] = attrs.get(f"authored_kind{suffix}", "") or ""
            row[doc_col] = attrs.get(f"document{suffix}", "") or ""

    def _fill_attribution_from_property_chain(self, row, pn) -> None:
        """Qualia claims: walk the provenance chain
        ``PropertyNode → has_data_provenance → Extractor|Combiner →
        (combines →) Extractor → has_author → Author`` and reconstruct
        the 1 or 2 attribution triples. Also handles the direct-author
        case (``PN → has_author → Author`` with no extractor).
        """
        # Direct has_author (author-only claim, no extractor step).
        direct_author = self._first_has_author(pn.node_id)
        # All has_data_provenance targets. At most one per PN in the
        # importer's output, but we tolerate a list defensively.
        prov_targets = []
        for edge in self.graph.edges:
            if (edge.edge_source == pn.node_id
                    and edge.edge_type == "has_data_provenance"):
                tgt = self.graph.find_node_by_id(edge.edge_target)
                if tgt is not None:
                    prov_targets.append(tgt)

        if not prov_targets:
            # No extractor chain; if the PN has a direct author, emit
            # a triple with empty EXTRACTOR / DOCUMENT.
            if direct_author is not None:
                self._write_triple(row, 0,
                                   extractor_text="",
                                   document_node=None,
                                   author_node=direct_author)
            return

        # If the provenance head is a Combiner, walk to both extractors.
        head = prov_targets[0]
        if isinstance(head, CombinerNode):
            row[self._COL_COMBINER] = getattr(head, "description", "") or ""
            extractors = []
            for edge in self.graph.edges:
                if (edge.edge_source == head.node_id
                        and edge.edge_type == "combines"):
                    ext = self.graph.find_node_by_id(edge.edge_target)
                    if isinstance(ext, ExtractorNode):
                        extractors.append(ext)
            for i, ext in enumerate(extractors[:2]):
                doc_node = self._extractor_document(ext)
                author = self._first_has_author(ext.node_id)
                self._write_triple(row, i,
                                   extractor_text=getattr(ext, "description", "") or "",
                                   document_node=doc_node,
                                   author_node=author)
            return

        # Single extractor (or unknown node type).
        if isinstance(head, ExtractorNode):
            doc_node = self._extractor_document(head)
            author = self._first_has_author(head.node_id)
            self._write_triple(row, 0,
                               extractor_text=getattr(head, "description", "") or "",
                               document_node=doc_node,
                               author_node=author)

    def _write_triple(self, row, idx, *,
                      extractor_text, document_node, author_node) -> None:
        """Write one attribution quadruple (EXTRACTOR_i / DOCUMENT_i /
        AUTHOR_i / AUTHOR_KIND_i) into ``row``. ``idx`` is 0 for the
        first triple, 1 for the second.
        """
        if idx == 0:
            ext_col, doc_col, auth_col, kind_col = (
                self._COL_EXT_1, self._COL_DOC_1,
                self._COL_AUTH_1, self._COL_KIND_1)
        else:
            ext_col, doc_col, auth_col, kind_col = (
                self._COL_EXT_2, self._COL_DOC_2,
                self._COL_AUTH_2, self._COL_KIND_2)

        row[ext_col] = extractor_text or ""
        if document_node is not None:
            row[doc_col] = self._doc_id_by_uuid.get(
                document_node.node_id, document_node.name or "")
        if author_node is not None:
            row[auth_col] = self._auth_id_by_uuid.get(
                author_node.node_id, author_node.name or "")
            row[kind_col] = (
                "extractor" if isinstance(author_node, AuthorAINode) else "author")

    def _first_has_author(self, origin_uuid):
        """First AuthorNode/AuthorAINode reachable from ``origin_uuid``
        via ``has_author``, or None.
        """
        for edge in self.graph.edges:
            if edge.edge_source != origin_uuid or edge.edge_type != "has_author":
                continue
            tgt = self.graph.find_node_by_id(edge.edge_target)
            if isinstance(tgt, AuthorNode):
                return tgt
        return None

    def _extractor_document(self, extractor_node):
        """First DocumentNode reachable from ``extractor_node`` via
        ``extracted_from``, or None.
        """
        for edge in self.graph.edges:
            if (edge.edge_source == extractor_node.node_id
                    and edge.edge_type == "extracted_from"):
                tgt = self.graph.find_node_by_id(edge.edge_target)
                if isinstance(tgt, DocumentNode):
                    return tgt
        return None


# ---------------------------------------------------------------------------
# Convenience top-level function
# ---------------------------------------------------------------------------

def write_unified_xlsx(graph: Graph, path: str) -> Dict[str, int]:
    """One-shot export: build a :class:`UnifiedXLSXExporter` and call
    its ``write`` method. Returns the row-count report.
    """
    return UnifiedXLSXExporter(graph).write(path)
