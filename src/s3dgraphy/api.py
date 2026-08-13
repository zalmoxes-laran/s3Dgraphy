"""s3dgraphy.api — the formal access-API surface (P1-F).

A single, documented set of **pure operations** over the Extended Matrix graph:
load/parse, validate, project → TTL/RDF, GraphML/XLSX interop, authority
resolution. This is the stable contract that the local **em-bridge** sidecar and
the future **em-server** (HTTP) both drive — the library itself stays pure:

  * NO web framework here (no FastAPI/uvicorn/HTTP). Transports wrap this surface.
  * Heavy/optional deps (lxml for GraphML, pandas for XLSX, rdflib for RDF) are
    imported LAZILY inside each function, so ``import s3dgraphy.api`` is cheap and
    ``pip install s3dgraphy`` pulls no web/format deps until an op needs them.
  * em.json is the single source of truth; ``dict`` = an em.json document,
    ``str``/``bytes`` = a serialized GraphML/Turtle payload, ``Graph`` = the
    in-memory model. HTTP callers work in docs/strings; in-process callers may
    use the Graph-level ops directly.

Layout is intentionally absent: it lives in em-core (Rust/WASM), not the Python
library (ADR-001 invariant 2).
"""

from __future__ import annotations

import math
import tempfile
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

# Type-only alias; avoid importing submodules at module load (keeps this import
# cheap and cycle-free). Real classes are imported lazily inside the ops.
Graph = Any
EmJson = Dict[str, Any]


class MissingDependency(ImportError):
    """An optional dependency for an op is not installed (e.g. rdflib for TTL).
    Transports may map this to a 501-style 'not available' response."""


# ── load / parse ──────────────────────────────────────────────────────────────
def _with_state_warnings(graph: Graph, warnings: List[str]) -> Tuple[Graph, List[str]]:
    """Refresh the graph's state warnings and fold them into the returned list.

    The em.json format carries no ``warnings`` section on purpose: warnings are
    a function of the graph's state, not a log of how it was loaded, so they are
    recomputed at every load instead of being persisted (F). Without this, an
    em.json opened from disk was silent about exactly the problems the GraphML
    path shouted about — the same graph, two different stories.

    Both channels are fed: ``graph.warnings`` for callers holding the graph, and
    the returned list for callers that only take the tuple (EMTools' import
    panel is one). The structured records land on ``graph.warning_records`` —
    see :func:`graph_warnings`.
    """
    from .edges.connection_resolver import recompute_warnings
    fresh = recompute_warnings(graph)
    merged = list(warnings)
    merged.extend(w for w in fresh if w not in warnings)
    return graph, merged


def graph_warnings(graph: Graph, *, recompute: bool = False
                   ) -> List[Dict[str, Any]]:
    """The graph's warnings as ``{kind, node_id, message}`` records.

    The load path already fills these, so the default is to read them; pass
    ``recompute=True`` after mutating the graph — the point of not persisting
    warnings is that they follow the state, so a UI can refresh them whenever it
    wants without an import round-trip.

    ``kind`` is one of ``connection_resolver.WARNING_KINDS``, the single spelling
    every consumer shares. ``node_id`` names the element to reveal when the
    reader clicks the warning (for an edge, its source). Edge records also carry
    ``edge_id``, ``target_id`` and the ``candidates`` the datamodel would allow.

    Warnings that are not state warnings — a stratigraphic cycle, a
    deserialisation note — have no record: this surface does not pretend to know
    what they point at. They remain in ``graph.warnings``.
    """
    from .edges.connection_resolver import (recompute_warnings,
                                            state_warning_records)
    if recompute:
        recompute_warnings(graph)
        return list(getattr(graph, "warning_records", []) or [])
    existing = getattr(graph, "warning_records", None)
    return list(existing) if existing else state_warning_records(graph)


def load_emjson(doc: EmJson) -> Tuple[Graph, List[str]]:
    """Parse an in-memory .em.json v1 document into a Graph.

    Returns ``(graph, warnings)``; unknown node types degrade to base nodes with
    a warning (forward-compatible) rather than failing. The state warnings
    (untyped nodes, role-less groups, degraded connections) are recomputed here
    — see :func:`_with_state_warnings`."""
    from .importer.emjson_importer import parse_emjson
    return _with_state_warnings(*parse_emjson(doc))


def load_emjson_file(path: str) -> Tuple[Graph, List[str]]:
    """Load a .em.json v1 file. Returns ``(graph, warnings)``, state warnings
    recomputed (F)."""
    from .importer.emjson_importer import import_emjson
    return _with_state_warnings(*import_emjson(path))


def graph_to_emjson(graph: Graph, layout: Optional[Dict[str, Any]] = None) -> EmJson:
    """Serialize a Graph to an .em.json v1 document (dict). Optional ``layout``
    is embedded as-is (the caller owns layout; em-core computes it)."""
    from .exporter.emjson_exporter import build_emjson
    return build_emjson(graph, layout=layout)


# ── validate ────────────────────────────────────────────────────────────────
def validate(graph: Graph) -> Dict[str, Any]:
    """Read-only structural check of a Graph. Returns
    ``{ok, stats, warnings, issues}`` — surfaces the graph's own accumulated
    warnings plus a cheap dangling-edge scan. Minimal and extensible; adds no
    side effects."""
    nodes = list(getattr(graph, "nodes", []) or [])
    edges = list(getattr(graph, "edges", []) or [])
    ids = {n.node_id for n in nodes}
    issues: List[str] = []
    for e in edges:
        if e.edge_source not in ids:
            issues.append(f"edge '{getattr(e, 'edge_id', '?')}' has missing source '{e.edge_source}'")
        if e.edge_target not in ids:
            issues.append(f"edge '{getattr(e, 'edge_id', '?')}' has missing target '{e.edge_target}'")
    warnings = list(getattr(graph, "warnings", []) or [])
    return {
        "ok": not issues,
        "stats": {"nodes": len(nodes), "edges": len(edges)},
        "warnings": warnings,
        "issues": issues,
    }


# ── the em.json CONTAINER: a project is one file ───────────────────────────────
#
# An em.json is ALWAYS a container: `{"graphs": {...}}`, 1..N study graphs plus
# the project shelf. A single graph is a container-of-one. This is the shape
# Heriverse already reads, so adopting it costs nobody a migration — and reading
# accepts the legacy single-graph form too, so nothing on disk breaks.

def load_container(doc: EmJson):
    """Read a container OR a legacy single-graph document → (Container, warnings)."""
    from .container import parse_container
    return parse_container(doc)


def load_container_file(path: str):
    """Read a container file (either shape) → (Container, warnings)."""
    from .container import load_container_file as _load
    return _load(path)


def container_to_emjson(container) -> EmJson:
    """Serialise a Container — always the `graphs` shape, one graph included."""
    from .container import build_container
    return build_container(container)


def save_container_file(container, path: str) -> str:
    from .container import save_container_file as _save
    return _save(container, path)


def container_of(graph: Graph, shelf: Optional[Graph] = None):
    """Wrap one graph (and optionally its shelf) as a container-of-one."""
    from .container import container_of as _wrap
    return _wrap(graph, shelf=shelf)


def merge_containers(container, other):
    """Integrate another project into this one — ADD its graphs, merge shared
    nodes by UUID. Returns a MergeReport.

    The offline "integrate later": no server, no session. Declared limit — this
    is add + merge-by-UUID, NOT conflict resolution; the report's `merged_nodes`
    is exactly the set where a divergent edit could have been overwritten.
    """
    from .container import merge_into_container
    return merge_into_container(container, other)


# ── 2D annotation → paradata chain ─────────────────────────────────────────────
def create_annotation_paradata(graph: Graph, image_id: str,
                               region: Dict[str, Any], interpretation: str,
                               property_type: str,
                               target_unit_id: Optional[str] = None,
                               author: Optional[str] = None):
    """One 2D annotation → the whole paradata chain (Extractor · Property ·
    AnnotationRegion, wired to the image and to the unit). Deterministic ids, so
    a re-send converges instead of duplicating.

    This is the seam the annotator's canvas will call — the semantics exist
    before any UI does, which is the point: a region drawn on a photograph is a
    CLAIM, and the chain is what makes it readable by somebody else.
    See :mod:`s3dgraphy.annotation`.
    """
    from .annotation import create_annotation_paradata as _create
    return _create(graph, image_id, region, interpretation, property_type,
                   target_unit_id=target_unit_id, author=author)


def create_geometry_proxy(graph: Graph, unit_id: str, shape: Dict[str, Any],
                          extractor_sources: Optional[List[str]] = None,
                          author: Optional[str] = None,
                          name: Optional[str] = None):
    """The PROXY of a unit, as a property with its provenance.

    The geometry-without-material of a US is a `PropertyNode(geometry)` whose
    payload is a SemanticShape (hulls/spheres inline or a `.glb`), not a lone
    SemanticShape hanging off the unit — so it inherits the paradata chain, and
    ONE proxy can be synthesised from several sources.
    See :mod:`s3dgraphy.geometry`.
    """
    from .geometry import create_geometry_proxy as _create
    return _create(graph, unit_id, shape, extractor_sources=extractor_sources,
                   author=author, name=name)


# ── project → TTL / RDF ────────────────────────────────────────────────────────
def project_ttl(graph: Graph, *, base_uri: Optional[str] = None,
                fmt: str = "turtle") -> str:
    """Project a Graph to RDF and return it as a string (default Turtle).

    Raises :class:`MissingDependency` if rdflib is not installed."""
    try:
        from .exporter.rdf_exporter import (
            export_single_graph_to_rdf, DEFAULT_BASE_URI,
        )
    except ImportError as exc:  # rdflib missing
        raise MissingDependency(f"RDF projection needs rdflib ({exc})") from exc
    with tempfile.NamedTemporaryFile(suffix=".ttl", delete=False) as tmp:
        tmp_path = Path(tmp.name)
    try:
        export_single_graph_to_rdf(
            graph, str(tmp_path), format=fmt,
            base_uri=base_uri or DEFAULT_BASE_URI)
        return tmp_path.read_text(encoding="utf-8")
    finally:
        tmp_path.unlink(missing_ok=True)


def emjson_to_ttl(doc: EmJson, *, base_uri: Optional[str] = None) -> str:
    """Convenience: load an em.json doc and project it to Turtle."""
    graph, _warnings = load_emjson(doc)
    return project_ttl(graph, base_uri=base_uri)


# ── GraphML interop ────────────────────────────────────────────────────────────
def graph_to_graphml(graph: Graph, *, persist_auxiliary: bool = False) -> str:
    """Serialize a Graph to yEd GraphML (XML string)."""
    from .exporter.graphml.graphml_exporter import GraphMLExporter
    with tempfile.NamedTemporaryFile(suffix=".graphml", delete=False) as tmp:
        tmp_path = Path(tmp.name)
    try:
        GraphMLExporter(graph).export(str(tmp_path), persist_auxiliary=persist_auxiliary)
        return tmp_path.read_text(encoding="utf-8")
    finally:
        tmp_path.unlink(missing_ok=True)


def graphml_to_graph(graphml: "str | bytes", *, graph_id: str = "imported_graph"
                     ) -> Tuple[Graph, List[str]]:
    """Parse yEd GraphML into a Graph. Returns ``(graph, warnings)``."""
    from .graph import Graph as _Graph
    from .importer.import_graphml import GraphMLImporter
    data = graphml.encode("utf-8") if isinstance(graphml, str) else graphml
    with tempfile.NamedTemporaryFile(suffix=".graphml", delete=False) as tmp:
        tmp.write(data)
        tmp_path = Path(tmp.name)
    try:
        g = _Graph(graph_id=graph_id)
        graph = GraphMLImporter(str(tmp_path), g).parse()
        return graph, list(getattr(graph, "warnings", []) or [])
    finally:
        tmp_path.unlink(missing_ok=True)


def emjson_to_graphml(doc: EmJson) -> str:
    """Convenience: load an em.json doc and serialize it to GraphML."""
    graph, _warnings = load_emjson(doc)
    return graph_to_graphml(graph)


def graphml_to_emjson(graphml: "str | bytes") -> EmJson:
    """Convenience: parse GraphML → em.json document (layout=None; the client
    re-lays-out via em-core)."""
    graph, _warnings = graphml_to_graph(graphml)
    return graph_to_emjson(graph)


# ── GraphML → em.json conversion (Se5) ────────────────────────────────────────
#: Fill colours that betray a graph drawn before the EM 1.4 palette. ``#CCCCFF``
#: is yEd's own default node colour: a node still wearing it was never given an
#: EM type by its author, and no amount of inference here can recover one. Used
#: ONLY to raise a hint — never to assign a type.
_LEGACY_EM_FILL_COLOURS = ("#CCCCFF",)


def _conversion_report(graph: Graph, warnings: List[str],
                       graphml: bytes) -> Dict[str, Any]:
    """Summarise what a GraphML → em.json conversion left unresolved.

    The counts are read off the GRAPH, not off the warning strings: a bare
    ``Node`` is what the importer produces when a yEd shape matches no EM type,
    a bare ``GroupNode`` when a box carries no palette colour, and
    ``generic_connection`` is the degraded edge. The raw warnings are carried
    along verbatim for display — they are the author-facing text (S6).
    """
    from .nodes.group_node import GroupNode
    from .nodes.stratigraphic_node import Node

    untyped = [n for n in graph.nodes if type(n) is Node]
    unclassified = [n for n in graph.nodes if type(n) is GroupNode]
    untyped_nodes = [getattr(n, "name", "") or n.node_id for n in untyped]
    unclassified_groups = [getattr(n, "name", "") or n.node_id
                           for n in unclassified]

    # A degraded edge is not automatically an authorial error. When one of its
    # endpoints has no type or no role, the edge COULD NOT have been resolved —
    # `handle_group_node` even writes `generic_connection` on purpose for a
    # role-less box. Those belong to the "author warning" bucket: fix the NODE
    # and the edge follows. What is left — both endpoints properly typed, yet
    # no relation the datamodel allows — is the real anomaly worth chasing.
    untypable_ids = {n.node_id for n in untyped} | {n.node_id for n in unclassified}
    degraded_edges = 0
    degraded_author_warning = 0
    for e in graph.edges:
        if getattr(e, "edge_type", "") != "generic_connection":
            continue
        degraded_edges += 1
        if (getattr(e, "edge_source", None) in untypable_ids
                or getattr(e, "edge_target", None) in untypable_ids):
            degraded_author_warning += 1
    degraded_real = degraded_edges - degraded_author_warning

    # Legacy-EM hint. Two signals must agree: the file still uses a pre-1.4
    # fill colour AND the importer could not type some nodes. Either alone is
    # too weak — a modern graph may reuse the colour decoratively, and an
    # untyped node may just be one bad shape.
    try:
        text = graphml.decode("utf-8", errors="replace")
    except Exception:  # pragma: no cover - defensive
        text = ""
    legacy_colours = [c for c in _LEGACY_EM_FILL_COLOURS if c in text]
    legacy = {
        "suspected": bool(legacy_colours and untyped_nodes),
        "evidence": ([f"fill colour {c} present" for c in legacy_colours]
                     + ([f"{len(untyped_nodes)} node(s) with no recognised EM type"]
                        if untyped_nodes else [])),
    }

    from .exporter.emjson_exporter import SCHEMA_VERSION
    return {
        "nodes": len(graph.nodes),
        "edges": len(graph.edges),
        "schema_version": SCHEMA_VERSION,
        "untyped_nodes": sorted(untyped_nodes),
        "unclassified_groups": sorted(unclassified_groups),
        "degraded_edges": degraded_edges,
        "degraded_edges_author_warning": degraded_author_warning,
        "degraded_edges_real": degraded_real,
        "legacy_em": legacy,
        "warnings": list(warnings),
    }


def convert_graphml_to_emjson(graphml: "str | bytes"
                              ) -> Tuple[EmJson, Dict[str, Any]]:
    """GraphML → ``(em.json document, conversion report)``.

    The 1.7 migration path. The GraphML is deprecated: 1.6 still imports it at
    runtime, but from 1.7 this one-shot conversion is the only way in. Because
    the importer mints deterministic ids (E2), the same GraphML always yields
    the same em.json — the conversion can therefore be treated as a persistent
    identity, not a throwaway.

    The report says what the conversion could NOT resolve (untyped nodes,
    unclassified groups, degraded edges) so the author can fix the SOURCE graph
    and convert again. Nothing is guessed here.
    """
    data = graphml.encode("utf-8") if isinstance(graphml, str) else graphml
    graph, warnings = graphml_to_graph(data)
    return graph_to_emjson(graph), _conversion_report(graph, warnings, data)


def format_conversion_report(report: Dict[str, Any], *,
                             max_warnings: int = 0,
                             list_warnings: bool = False) -> str:
    """Human-readable rendering of :func:`convert_graphml_to_emjson`'s report.

    ``max_warnings`` truncates the warning list (0 = all); ``list_warnings``
    prints the individual warning lines rather than just the aggregate counts.
    """
    lines = [
        f"converted: {report['nodes']} nodes, {report['edges']} edges "
        f"(em.json schema_version {report['schema_version']})",
    ]

    untyped = report["untyped_nodes"]
    groups = report["unclassified_groups"]
    degraded = report["degraded_edges"]

    if not (untyped or groups or degraded):
        lines.append("nothing left unresolved.")
    else:
        lines.append("")
        lines.append("left unresolved — fix these in the SOURCE graph, not here:")
        if untyped:
            lines.append(f"  {len(untyped)} node(s) with no recognised EM type "
                         f"(yEd shape/colour matches nothing): "
                         f"{', '.join(untyped[:10])}"
                         + (" …" if len(untyped) > 10 else ""))
        if groups:
            lines.append(f"  {len(groups)} group(s) with no EM role "
                         f"(no palette colour): {', '.join(groups[:10])}"
                         + (" …" if len(groups) > 10 else ""))
        if degraded:
            real = report.get("degraded_edges_real", degraded)
            follow = report.get("degraded_edges_author_warning", 0)
            lines.append(f"  {degraded} edge(s) degraded to generic_connection")
            if follow:
                lines.append(f"      {follow} of them touch an untyped node or "
                             f"a role-less group — fix the NODE and the edge "
                             f"follows")
            if real:
                lines.append(f"      {real} with both endpoints typed: real "
                             f"authorial anomalies, worth chasing one by one")

    if report["legacy_em"]["suspected"]:
        lines.append("")
        lines.append("⚠ this looks like a pre-1.4 EM graph:")
        for e in report["legacy_em"]["evidence"]:
            lines.append(f"    - {e}")
        lines.append("  Run the EMTools operator «convert EM 1.x → 1.4» on the "
                     "GraphML FIRST, then convert again. The types are not "
                     "guessed here: an ambiguous colour has no single reading.")

    warnings = report["warnings"]
    if warnings:
        lines.append("")
        lines.append(f"{len(warnings)} importer warning(s).")
        if list_warnings:
            shown = warnings if max_warnings <= 0 else warnings[:max_warnings]
            for w in shown:
                lines.append(f"  - {w}")
            if len(shown) < len(warnings):
                lines.append(f"  … {len(warnings) - len(shown)} more "
                             f"(raise --max-warnings, or 0 for all)")
        else:
            lines.append("  (re-run with --list-warnings to see them)")

    return "\n".join(lines)


# ── EM Narrative — generation seam (N5) ───────────────────────────────────────
#
# s3Dgraphy stays PURE: these two build the briefing and write the result back,
# and neither touches the network. The model call itself lives in em-bridge,
# behind a provider interface — the same split that keeps this package free of
# web frameworks.


def build_narrative_generation_context(graph: Graph, activity_id: str = None, *,
                                       chapter_ref: str = None,
                                       template_id: str = "site_story"
                                       ) -> Dict[str, Any]:
    """Everything a model needs to write about one activity, as a plain dict.

    The activity, its actions in stratigraphic order, their epochs, the evidence
    chain behind each, the sources that chain rests on, and the style contract.
    Deliberately a briefing and not a dump: what is not in it cannot leak.
    """
    from .narrative.generation import build_narrative_generation_context as _b
    return _b(graph, activity_id, chapter_ref=chapter_ref,
              template_id=template_id)


def write_ai_draft(graph: Graph, target: str, text: str, *, model: str,
                   version: str = "", date: Optional[str] = None,
                   prompt: str = "", narrative_id: Optional[str] = None,
                   chapter_title: Optional[str] = None) -> Dict[str, Any]:
    """Write generated prose into the narrative: attributed to the AI author,
    with the prompt registered as a source, and **unendorsed**.

    Nothing here validates anything — validation is an act by a person (N4)."""
    from .narrative.generation import write_ai_draft as _w
    return _w(graph, target, text, model=model, version=version, date=date,
              prompt=prompt, narrative_id=narrative_id,
              chapter_title=chapter_title)


# ── EM Narrative — print projection (L1) ──────────────────────────────────────
def export_narrative_latex(graph: Graph, narrative_id: str) -> Dict[str, str]:
    """Project a NarrativeNode to LaTeX + BibTeX: ``{"tex": …, "bib": …}``.

    An **exporter, not a renderer**: it returns two strings and needs no LaTeX
    engine. ``tex`` is a body to ``\\input{}`` into your own preamble (chapters →
    ``\\section``/``\\subsection``, prose → prose, embeds → figures or
    ``\\cite``), ``bib`` one entry per cited source with keys derived
    deterministically from the node ids (:func:`bib_key`), so citations and
    entries cannot drift and a re-export is stable.

    Which embeds become citations rather than figures is decided by their
    ``view_type``: `source` and `document` are things a reader could go and read;
    everything else is something they look at. Nothing is invented — a source
    without metadata becomes a minimal ``@misc`` with the title it has.

    Raises ``KeyError`` if ``narrative_id`` names no narrative in the graph."""
    from .exporter.latex_exporter import export_narrative_latex as _e
    return _e(graph, narrative_id)


def bib_key(node_id: str) -> str:
    """The stable BibTeX key for a node id (``em:<slug>``) — exposed so a caller
    can cite an EM source from its own document without re-deriving the rule."""
    from .exporter.latex_exporter import bib_key as _k
    return _k(node_id)


def bake_narrative(graph: Graph, narrative_id: str, *,
                   base_dir: Optional[str] = None) -> Any:
    """Resolve a live narrative into a **static snapshot** (``BakedNarrative``).

    A narrative's embeds mean "whatever this node says now" — which is what makes
    it an editing surface and what a published text cannot be. The bake commits to
    one reading, once: citations resolved, image bytes read, coordinates converted,
    and a labelled placeholder wherever a static form needs a renderer this build
    does not have (3D scene, matrix).

    **Format-agnostic on purpose.** DocX, LaTeX, HTML and a notebook all render
    the same bake, so they cannot disagree about what the narrative said — three
    separate traversals of the graph could, and would.

    Nothing raises over a missing file and nothing is dropped: an embed that will
    not resolve becomes a block that says so, and is also listed in
    ``baked.unresolved`` — the list to check before publishing. A snapshot may
    record a hole; it may not hide one.

    ``base_dir`` is what relative image locators resolve against (normally the
    folder holding the em.json). Raises ``KeyError`` if ``narrative_id`` names no
    narrative.
    """
    from .narrative.bake import bake_narrative as _b
    return _b(graph, narrative_id, base_dir=base_dir)


def export_narrative_docx(graph: Graph, narrative_id: str, *,
                          base_dir: Optional[str] = None) -> bytes:
    """Render a narrative to **.docx** bytes — the format for the normal reader.

    Bakes first (:func:`bake_narrative`), then places the result: chapter
    headings, prose, embedded images, citations as text plus a "Fonti" section,
    and the byline that keeps people (responsible) apart from models (assisting).
    Unendorsed machine drafts are flagged **in the text**, because a Word file gets
    printed, copied and re-saved, and each of those loses anything that is not a
    character.

    Returns bytes rather than writing a file: the caller may be a CLI, an HTTP
    response or a notebook.

    Raises :class:`MissingDependency` if python-docx is not installed — the bake
    itself does not need it, so ``bake_narrative`` keeps working and only the
    rendering is unavailable.
    """
    from .exporter.docx_exporter import render_docx
    return render_docx(bake_narrative(graph, narrative_id, base_dir=base_dir))


# ── XLSX mapping ──────────────────────────────────────────────────────────────
def xlsx_to_graph(path: str, *, mapping_name: Optional[str] = None,
                  id_column: Optional[str] = None, graph_id: str = "imported_graph"
                  ) -> Tuple[Graph, List[str]]:
    """Map an Excel workbook into a Graph via the XLSX importer. Returns
    ``(graph, warnings)``. Needs pandas/openpyxl (imported lazily by the
    importer)."""
    from .graph import Graph as _Graph
    from .importer.xlsx_importer import XLSXImporter
    importer = XLSXImporter(path, mapping_name=mapping_name, id_column=id_column)
    # BaseImporter subclasses attach to a Graph they hold; use the importer's own.
    graph = importer.parse()
    if getattr(graph, "graph_id", None) in (None, ""):
        graph.graph_id = graph_id
    return graph, list(getattr(graph, "warnings", []) or [])


def xlsx_to_emjson(path: str, **kwargs) -> EmJson:
    """Convenience: map an Excel workbook → em.json document."""
    graph, _warnings = xlsx_to_graph(path, **kwargs)
    return graph_to_emjson(graph)


# ── StratiMiner — assisted graph creation from unstructured sources ──────────
#
# The pipeline is deliberately in two halves, and the seam between them is a
# FILE a human can open:
#
#     a folder of documents  --(AI)-->  em_data.xlsx  --(deterministic)-->  em.json
#
# The AI writes ONLY the middle artefact. It never writes the graph. That is
# the whole point of having an intermediate table rather than asking a model
# for em.json directly: canonisation stays reviewable *before* it becomes
# structure, and the second arrow is a plain importer that anyone can re-run
# and get the same bytes. A model that emitted em.json would put an unreviewed
# guess straight into the language's own format, where a wrong node type reads
# exactly like a right one.
#
# Both halves live here, in the pure library, so neither depends on Blender or
# on a UI: EMtools, EMStudio (via em-bridge) and a bare script call the same
# two functions.


def em_data_to_graph(path: str, *, graph_id: Optional[str] = None
                     ) -> Tuple[Graph, List[str], Dict[str, int]]:
    """Read a canonical ``em_data.xlsx`` (5 typed sheets) into a Graph.

    Returns ``(graph, warnings, stats)``. Needs pandas/openpyxl, imported
    lazily by the importer.

    **Not the same thing as :func:`xlsx_to_graph`.** That one maps an arbitrary
    workbook through a *named mapping* (pyArchInit and friends); this one reads
    the ONE canonical shape StratiMiner produces — sheets ``Units``, ``Epochs``,
    ``Claims``, ``Authors``, ``Documents``. Folding them into a single entry
    point would mean sniffing the file to decide which contract applies, and a
    wrong guess there fails deep inside a parser instead of at the door.

    **Why ``stats`` rides along** (the one place this deviates from
    ``xlsx_to_graph``'s two-tuple): the intermediate table exists in order to be
    checked, so the caller has to be able to say "5 units, 2 epochs, 7 claims
    read → 14 nodes, 9 edges" *before* anybody trusts the graph. Recomputing
    that from the finished graph cannot distinguish a row that was skipped from
    a row that never existed, and re-opening the workbook to count again is a
    second parse free to disagree with the first.
    """
    from .importer.unified_xlsx_importer import UnifiedXLSXImporter
    importer = UnifiedXLSXImporter(path, graph_id=graph_id)
    graph = importer.parse()
    return (graph,
            list(getattr(graph, "warnings", []) or []),
            dict(importer.stats))


def import_em_data(path: str, *, graph_id: Optional[str] = None) -> EmJson:
    """``em_data.xlsx`` → em.json document. The deterministic half of
    StratiMiner: same workbook in, same document out, no model involved.

    Use :func:`em_data_to_graph` when you also want the warnings and the row
    counts (a UI does).
    """
    graph, _warnings, _stats = em_data_to_graph(path, graph_id=graph_id)
    return graph_to_emjson(graph)


def stratiminer_prompt(*, language: Optional[str] = None,
                       documents_folder: Optional[str] = None,
                       document_list: Optional[List[Any]] = None,
                       include_validation: bool = True,
                       include_checklist: bool = True,
                       include_stratigraphy_only: bool = False,
                       dosco_in_place: bool = True,
                       ai_has_filesystem_access: bool = True) -> str:
    """Build the StratiMiner extraction prompt for a folder of sources.

    A pure string builder: it reads the prompt template bundled in the package
    and substitutes the caller's options. No model is called here — the result
    is what you either hand to a Cowork session (the user runs it) or send
    through a provider seam (the host application does).

    Re-exported on ``api`` although the implementation lives in ``utils``,
    because ``api`` is the consumption surface (ADR-001): a caller outside
    Blender should not have to know which module the template loader sits in,
    and the name should say what the prompt is *for*.

    The contract of the returned prompt is the invariant above: it asks the
    model for **em_data.xlsx**, never for a graph.
    """
    from .utils.utils import get_ai_prompt
    return get_ai_prompt(
        language=language,
        include_validation=include_validation,
        include_checklist=include_checklist,
        include_stratigraphy_only=include_stratigraphy_only,
        documents_folder=documents_folder,
        document_list=document_list,
        dosco_in_place=dosco_in_place,
        ai_has_filesystem_access=ai_has_filesystem_access,
    )


def source_text(path: str, *, max_chars: Optional[int] = None
                ) -> Dict[str, Any]:
    """What a source document says, for Path A's prompt — or why it cannot be read.

    ``{"text": str | None, "kind": "text"|"pdf"|"unsupported", "note": str}``.
    Never raises: a caller cataloguing a folder must not lose nineteen documents to
    one bad twentieth. When ``text`` is ``None`` the ``note`` says why, and that
    note is what the prompt shows the model so it does not invent the contents of a
    file nobody read.

    PDFs need the ``[pdf]`` extra (pypdf, ~350 KB, pure Python — **not** PyMuPDF's
    21 MB). Without it a PDF reports that its text was not read and StratiMiner
    degrades to filenames, which is a valid request this build cannot serve, not an
    error.
    """
    from .importer.source_text import DEFAULT_MAX_CHARS, source_text as _st
    return _st(path, max_chars=DEFAULT_MAX_CHARS if max_chars is None
               else max_chars)


def pdf_text_available() -> bool:
    """True when this build can read a PDF's text layer. Ask ONCE, before the work,
    so a folder of scans produces one honest sentence instead of twenty notes."""
    from .importer.source_text import pdf_text_available as _a
    return _a()


def source_text_extractor() -> Optional[str]:
    """The PDF extractor in use, with its version, or ``None``. Recorded because
    two extractors disagree about hyphenation and column order: a canonisation that
    came out oddly should be traceable to whatever read the page."""
    from .importer.source_text import extractor_name as _n
    return _n()


def em_data_sheets() -> Tuple[str, ...]:
    """The sheet names a canonical ``em_data.xlsx`` must carry, in the order the
    importer reads them. Exposed so a UI or a writer (the Path-A materialiser in
    em-bridge) enumerates them from here instead of hardcoding five strings that
    then drift from the importer."""
    from .importer.unified_xlsx_importer import UnifiedXLSXImporter
    return tuple(UnifiedXLSXImporter._SHEETS)


def em_data_columns() -> Dict[str, Tuple[str, ...]]:
    """The column layout of each sheet, for a caller that has to WRITE the
    table. Same source as :func:`em_data_sheets`."""
    from .importer.unified_xlsx_importer import UnifiedXLSXImporter
    return {k: tuple(v) for k, v in UnifiedXLSXImporter._COLUMNS.items()}


def write_em_data(sheets: Dict[str, List[Dict[str, Any]]], path: str
                  ) -> Dict[str, Any]:
    """Write a canonical ``em_data.xlsx`` from row dicts. Returns a report:
    ``{"path", "rows": {sheet: n}, "warnings": [...]}``.

    This is the **materialiser** of StratiMiner's first arrow. A language model
    cannot hand back a binary workbook — it returns rows — so somebody has to
    turn rows into the file, and that somebody belongs here rather than in the
    caller: the header layout lives next to the importer that reads it, and
    EMtools, em-bridge and a bare script then produce byte-comparable tables
    instead of three nearly-identical writers.

    **Unknown column names are dropped, with a warning, not written.** This is
    the guard that keeps the seam honest: what arrives here came from a model,
    and a model that invents a column would otherwise get it silently carried
    into the workbook, where the importer ignores it — so the information would
    appear to have been captured while being nowhere. A named warning turns that
    into something the user can see and fix in the table, which is exactly what
    the table is for.

    Needs openpyxl (imported lazily).
    """
    from openpyxl import Workbook
    from .importer.unified_xlsx_importer import UnifiedXLSXImporter

    columns = UnifiedXLSXImporter._COLUMNS
    warnings: List[str] = []
    rows_written: Dict[str, int] = {}

    unknown_sheets = [s for s in sheets if s not in columns]
    for s in sorted(unknown_sheets):
        warnings.append(
            f"sheet '{s}' is not part of em_data.xlsx "
            f"({', '.join(UnifiedXLSXImporter._SHEETS)}); not written")

    wb = Workbook()
    wb.remove(wb.active)
    for name in UnifiedXLSXImporter._WRITE_ORDER:
        header = columns[name]
        ws = wb.create_sheet(name)
        ws.append(list(header))
        for index, row in enumerate(sheets.get(name) or [], start=2):
            if not isinstance(row, dict):
                warnings.append(
                    f"{name} row {index}: expected an object of "
                    f"column->value, got {type(row).__name__}; skipped")
                continue
            for key in row:
                if key not in header:
                    warnings.append(
                        f"{name} row {index}: unknown column '{key}' "
                        f"dropped (columns: {', '.join(header)})")
            ws.append([_cell(row.get(col)) for col in header])
        rows_written[name] = max(ws.max_row - 1, 0)

    wb.save(path)
    return {"path": path, "rows": rows_written, "warnings": warnings}


def _cell(value: Any) -> Any:
    """Flatten a JSON value into something a spreadsheet cell can hold.

    openpyxl refuses lists and dicts, and a model does sometimes return
    ``["A.01", "A.02"]`` where the schema says a comma-separated string. Joining
    is the reading the importer already expects for the multi-value columns
    (``AUTHOR_IDS``), so accept it rather than failing the whole workbook over
    a formatting habit.
    """
    if value is None:
        return ""
    if isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, (list, tuple)):
        return ", ".join(str(v) for v in value if v is not None)
    return str(value)


# ── authority resolution ────────────────────────────────────────────────────
def resolve_authority(term: str, facet: str) -> List[Dict[str, Any]]:
    """Ranked offline authority candidates for a term/facet (P1-D). Returns [] on
    empty term or when the resolver/snapshots are unavailable."""
    try:
        from .authorities import resolve
        return list(resolve(term, facet))
    except Exception:
        return []


def authority_facets() -> List[str]:
    """The valid authority facet identifiers (WHEN/WHAT/WHERE/WHO)."""
    try:
        from .authorities import FACET_ORDER
        return list(FACET_ORDER)
    except Exception:
        return []


# ── coordinate reprojection (G1) ──────────────────────────────────────────────
# Excavation coordinates are normally PROJECTED — a UTM zone, a national grid —
# with the EPSG code recorded next to them on the graph's GeoPositionNode. A web
# map wants WGS84 degrees. Converting between the two is not arithmetic anyone
# should improvise: it needs PROJ, and it belongs here, once, rather than in
# every consumer (and emphatically not in TypeScript, where a wrong guess would
# put a site in the Gulf of Guinea and look authoritative doing it).
#
# pyproj is the dependency and it is OPTIONAL and LAZY, exactly like rdflib for
# TTL: importing s3dgraphy stays free, and a build without the [geo] extra raises
# MissingDependency at the one op that needs it. pyproj rather than GDAL because
# its wheels bundle PROJ — pip installs it, no system libraries.
def reproject(x: float, y: float, epsg_source: int,
              epsg_target: int = 4326) -> Tuple[float, float]:
    """Convert one coordinate pair between two EPSG frames.

    Returns ``(lon, lat)`` in degrees when ``epsg_target`` is 4326 (the default),
    otherwise ``(x, y)`` in the target frame's own units. Axis order is always
    **easting/northing → lon/lat**: pyproj is asked for ``always_xy=True``, so
    callers never have to know that EPSG:4326 formally declares lat before lon —
    the classic way to end up with the coordinates swapped.

    Identity is short-circuited: ``epsg_source == epsg_target`` returns the input
    untouched and needs no pyproj at all, so a graph already in WGS84 works in a
    build without the [geo] extra.

    Raises :class:`MissingDependency` if pyproj is not installed, and
    ``ValueError`` for an unusable EPSG or a non-finite result (PROJ signals a
    point outside the source frame's domain with infinities — a silent ``inf``
    would travel all the way to a marker somewhere absurd).
    """
    xs, ys = reproject_many([(x, y)], epsg_source, epsg_target)[0]
    return xs, ys


def reproject_many(points: List[Tuple[float, float]], epsg_source: int,
                   epsg_target: int = 4326) -> List[Tuple[float, float]]:
    """:func:`reproject` for many points, building the transformer ONCE.

    A footprint is four corners plus a centroid; five transformer constructions
    to convert five points would be five PROJ pipeline lookups for one answer.
    """
    try:
        src = int(epsg_source)
        dst = int(epsg_target)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"EPSG codes must be integers: {exc}") from exc
    pts = [(float(px), float(py)) for px, py in points]
    if src == dst:
        return pts
    try:
        from pyproj import Transformer
    except ImportError as exc:  # pyproj missing
        raise MissingDependency(
            f"coordinate reprojection needs pyproj — install the [geo] extra "
            f"({exc})") from exc
    try:
        transformer = Transformer.from_crs(
            f"EPSG:{src}", f"EPSG:{dst}", always_xy=True)
    except Exception as exc:  # unknown/unsupported EPSG
        raise ValueError(
            f"cannot build a transformer EPSG:{src} → EPSG:{dst}: {exc}") from exc
    out: List[Tuple[float, float]] = []
    for px, py in pts:
        rx, ry = transformer.transform(px, py)
        if not (math.isfinite(rx) and math.isfinite(ry)):
            raise ValueError(
                f"({px}, {py}) is outside the domain of EPSG:{src} "
                f"(PROJ returned a non-finite result)")
        out.append((rx, ry))
    return out


# ── georeferencing a whole scene (G3) ─────────────────────────────────────────
# G1 answers "where is this point, in degrees". G3 answers the question a reader
# actually asks of a map: **how does the scene sit on the ground** — where it is
# AND which way it faces.
#
# The transform is the graph's own statement, in three steps and one order:
#
#     rotate by the azimuth  →  add the origin (shift)  →  reproject epsg → 4326
#
# and the order is not negotiable: rotating after translating would swing the
# scene around the grid origin instead of around itself, which for a shift a few
# hundred metres away puts the building in the next field. The azimuth is
# `rotation` on the graph-level GeoPositionNode (G1), clockwise from north, and 0
# — north up — must be the identity.
def _geo_anchor(graph: Graph) -> Dict[str, Any]:
    """The graph's georeferencing anchor as a plain dict, defaults included.

    One per graph by construction (``Graph.__init__``); if a document carries
    several, the one with the canonical id ``geo_<graph_id>`` wins and otherwise
    the first found — a graph with two anchors is a data problem, not something to
    average.
    """
    nodes = [n for n in (getattr(graph, "nodes", []) or [])
             if getattr(n, "node_type", None) == "geo_position"]
    if not nodes:
        return {"epsg": 4326, "shift_x": 0.0, "shift_y": 0.0, "shift_z": 0.0,
                "rotation": 0.0}
    canonical = f"geo_{getattr(graph, 'graph_id', '')}"
    node = next((n for n in nodes if getattr(n, "node_id", None) == canonical),
                nodes[0])
    data = dict(getattr(node, "data", {}) or {})
    return {
        "epsg": int(data.get("epsg") or 4326),
        "shift_x": float(data.get("shift_x") or 0.0),
        "shift_y": float(data.get("shift_y") or 0.0),
        "shift_z": float(data.get("shift_z") or 0.0),
        "rotation": float(data.get("rotation") or 0.0),
    }


#: EPSG codes that are GEOGRAPHIC (degrees), used when pyproj cannot be asked.
#: Short on purpose: it only has to cover what an EM graph realistically declares
#: as its anchor, and with pyproj installed the authoritative answer is used.
_GEOGRAPHIC_EPSG = {4326, 4979, 4258, 4269, 4230, 4267}


def _is_geographic(epsg: int) -> bool:
    """True when the frame's units are DEGREES rather than metres."""
    try:
        from pyproj import CRS
        return bool(CRS.from_epsg(int(epsg)).is_geographic)
    except Exception:
        return int(epsg) in _GEOGRAPHIC_EPSG


def georeference_scene(graph: Graph, points_local: List[Tuple[float, float]], *,
                       epsg_target: int = 4326) -> Dict[str, Any]:
    """Place scene-local points on the earth: rotate → shift → reproject.

    ``points_local`` are XY in the scene's own frame (metres, origin at the
    scene's 0,0) — typically the four corners of a bounding box and its centroid.
    Returns::

        {"epsg_source": int, "epsg_target": int, "rotation": float,
         "shift": [x, y, z], "points": [[lon, lat], …], "reprojected": bool}

    ``reprojected`` says whether PROJ was actually needed: with an anchor already
    in the target frame the points are exact without pyproj, and a caller can
    report that honestly rather than implying a conversion that did not happen.

    Raises :class:`MissingDependency` when the anchor's frame needs pyproj and it
    is not installed, and ``ValueError`` for an unusable EPSG — the same contract
    as :func:`reproject`, because this IS that op with the scene's pose applied
    first. Nothing here invents geometry: if you have no points, you get none.
    """
    anchor = _geo_anchor(graph)
    # A scene-local extent is METRES. Adding metres to an anchor expressed in
    # DEGREES is not a small inaccuracy, it is a category error: 30 m would become
    # 30 degrees and the footprint would span a continent. Refuse, and say what
    # the graph would need — a projected CRS on its GeoPositionNode.
    if any(px or py for px, py in points_local) and _is_geographic(anchor["epsg"]):
        raise ValueError(
            f"the graph's georeferencing anchor is in EPSG:{anchor['epsg']}, "
            f"which is in degrees: a scene extent in metres cannot be composed "
            f"with it. Give the GeoPositionNode a projected CRS (a UTM zone, a "
            f"national grid) — that is what an excavation records anyway.")
    theta = math.radians(anchor["rotation"])
    cos_t, sin_t = math.cos(theta), math.sin(theta)
    placed: List[Tuple[float, float]] = []
    for px, py in points_local:
        x, y = float(px), float(py)
        # Clockwise from north: a scene rotated +90° has its local +Y pointing
        # east. With rotation 0 this is exactly the identity (cos 0 = 1, sin 0 = 0),
        # which is the property the test pins.
        rx = x * cos_t + y * sin_t
        ry = -x * sin_t + y * cos_t
        placed.append((rx + anchor["shift_x"], ry + anchor["shift_y"]))
    src = anchor["epsg"]
    out = reproject_many(placed, src, epsg_target) if placed else []
    return {
        "epsg_source": src,
        "epsg_target": int(epsg_target),
        "rotation": anchor["rotation"],
        "shift": [anchor["shift_x"], anchor["shift_y"], anchor["shift_z"]],
        "points": [[x, y] for x, y in out],
        "reprojected": src != int(epsg_target),
    }


def scene_extent(graph: Graph) -> Optional[Dict[str, Any]]:
    """The scene's local XY extent, DERIVED from geometry the graph already has.

    The only geometry an EM graph holds is the spatial proxies:
    :class:`~s3dgraphy.nodes.semantic_shape_node.SemanticShapeNode` carries
    ``convexshapes`` (flat ``[x,y,z, x,y,z, …]`` vertex lists) and ``spheres``
    (``[x,y,z,r]``). Where those exist, an extent is a fact and this returns it::

        {"min_x": …, "min_y": …, "max_x": …, "max_y": …,
         "centroid": [x, y], "corners": [[x,y] × 4], "source": "semantic_shape"}

    Where they do not, this returns ``None`` — and a caller must then be given the
    extent explicitly or draw nothing. **A bounding box nobody measured is not a
    default, it is a fabrication**, and a fabricated footprint on a map is exactly
    the kind of confident wrong answer this codebase refuses elsewhere.
    """
    min_x = min_y = float("inf")
    max_x = max_y = float("-inf")
    seen = False
    for node in getattr(graph, "nodes", []) or []:
        if getattr(node, "node_type", None) != "semantic_shape":
            continue
        data = getattr(node, "data", {}) or {}
        for verts in (data.get("convexshapes") or []):
            for i in range(0, len(verts) - 2, 3):
                x, y = float(verts[i]), float(verts[i + 1])
                min_x, max_x = min(min_x, x), max(max_x, x)
                min_y, max_y = min(min_y, y), max(max_y, y)
                seen = True
        for sphere in (data.get("spheres") or []):
            if len(sphere) < 4:
                continue
            x, y, _z, r = (float(sphere[0]), float(sphere[1]),
                           float(sphere[2]), abs(float(sphere[3])))
            min_x, max_x = min(min_x, x - r), max(max_x, x + r)
            min_y, max_y = min(min_y, y - r), max(max_y, y + r)
            seen = True
    if not seen:
        return None
    return {
        "min_x": min_x, "min_y": min_y, "max_x": max_x, "max_y": max_y,
        "centroid": [(min_x + max_x) / 2.0, (min_y + max_y) / 2.0],
        # Corner order is fixed and documented: SW, SE, NE, NW in the scene's own
        # frame, so a consumer can draw a closed ring without guessing a winding.
        "corners": [[min_x, min_y], [max_x, min_y], [max_x, max_y], [min_x, max_y]],
        "source": "semantic_shape",
    }


# ── resource layer (R0: stable-ID resolver seam) ──────────────────────────────
# Resources are LinkNodes (E73). Their stable, storage-agnostic ID is the node
# UUID; the ``url`` is the *current locator*, not the identity. A pluggable
# resolver maps ID → a concrete Location. See :mod:`s3dgraphy.resources`.
def _resource_locator(node: Any) -> str:
    """The current locator of a resource node (ResourceNode ``url``)."""
    url = getattr(node, "url", None)
    if url is None:
        url = (getattr(node, "data", {}) or {}).get("url", "")
    return url or ""


def list_resources(graph: Graph) -> List[Dict[str, Any]]:
    """List the resources in a graph. Resources = LinkNodes (E73). Returns, per
    resource, its stable ``id`` (node UUID), ``name``, current ``locator`` and
    the classified location ``kind`` — a pure read, no I/O on the locator."""
    from .resources import classify_locator, stable_resource_id
    out: List[Dict[str, Any]] = []
    for node in getattr(graph, "nodes", []) or []:
        if getattr(node, "node_type", None) != "resource":
            continue
        locator = _resource_locator(node)
        out.append({
            "id": stable_resource_id(node),
            "name": getattr(node, "name", ""),
            "locator": locator,
            "kind": classify_locator(locator),
        })
    return out


def shelf_resources(doc: Optional[EmJson], folder: str, *,
                    graph_code: Optional[str] = None) -> List[Dict[str, Any]]:
    """The **Shelf** for a library/DosCo ``folder`` given the current em.json
    ``doc``: the un-hatted resources (orphans) with their stable IDs.

    Scans ``folder`` with the FS-index backend (R1), then returns the orphans —
    files with no matching graph node (DosCo EM-id convention filters applied) —
    additionally excluding any whose FS stable ID is already a node in ``doc``
    (ID-adoption aware, so a hatted resource leaves the Shelf). ``doc`` may be
    ``None`` (empty graph → every on-convention Document-id file is on the Shelf).
    Returns dicts ``{resource_id, key_id, filename, rel_path}``.

    Manifest-backed: reuses a ``.em_resources_manifest.json`` persisted IN the
    folder (same filename EMTools R4 uses) so stable IDs survive across scans and
    across tools — one ID space per folder, the "single connector". A hatted
    resource therefore keeps the same ID that was adopted as its node_id."""
    import json as _json
    import os as _os
    from .resources import FSIndexBackend
    manifest_name = ".em_resources_manifest.json"
    mpath = _os.path.join(folder, manifest_name)
    backend = None
    if _os.path.isfile(mpath):
        try:
            with open(mpath, encoding="utf-8") as fh:
                backend = FSIndexBackend.from_manifest(_json.load(fh))
            backend.folder = _os.path.abspath(folder)
        except Exception:
            backend = None
    if backend is None:
        backend = FSIndexBackend(folder)
    backend.rescan()
    if _os.path.isdir(folder):
        try:
            with open(mpath, "w", encoding="utf-8") as fh:
                _json.dump(backend.to_manifest(), fh, indent=2)
        except Exception:
            pass

    # The Shelf needs only each node's id/name/node_type — read them straight
    # from the em.json dict rather than a full ``load_emjson`` parse, so the
    # Shelf stays robust to header/version quirks (e.g. a freshly-created empty
    # doc) and stays decoupled from the importer. A tiny shim gives
    # ``FSIndexBackend.orphans`` the node view it expects (name-convention
    # exclusion); the id-adoption exclusion uses the same ids.
    class _ShimNode:
        __slots__ = ("node_id", "name", "node_type")

        def __init__(self, nid, name, ntype):
            self.node_id, self.name, self.node_type = nid, name, ntype

    class _ShimGraph:
        def __init__(self, nodes):
            self.nodes = nodes

    graph = None
    node_ids: set = set()
    if doc is not None:
        raw_nodes = ((doc.get("graph") or {}).get("nodes")) or []
        shim = [_ShimNode(n.get("id"), n.get("name", ""), n.get("node_type") or n.get("type"))
                for n in raw_nodes if isinstance(n, dict)]
        graph = _ShimGraph(shim)
        node_ids = {n.node_id for n in shim if n.node_id}
    out: List[Dict[str, Any]] = []
    for orphan in backend.orphans(graph, graph_code=graph_code):
        if orphan.resource_id in node_ids:
            continue  # already hatted (ID adopted) — off the Shelf
        out.append({
            "resource_id": orphan.resource_id,
            "key_id": orphan.key_id,
            "filename": orphan.filename,
            "rel_path": orphan.rel_path,
        })
    return out


def resolve_resource(graph: Graph, resource_id: str, *, registry: Any = None
                     ) -> Optional[Dict[str, Any]]:
    """Resolve a resource's stable ID to a Location dict via the resolver.

    Looks up the ResourceNode by its ``node_id`` (the stable ID), reads its current
    locator, and asks the resolver registry (default: passthrough) to map it to
    a :class:`~s3dgraphy.resources.Location`. Returns the Location as a dict
    ``{kind, value, exists}``, or ``None`` if no such resource exists."""
    from .resources import default_registry
    node = graph.find_node_by_id(resource_id)
    if node is None or getattr(node, "node_type", None) != "resource":
        return None
    locator = _resource_locator(node)
    reg = registry if registry is not None else default_registry()
    loc = reg.resolve(resource_id, locator, graph=graph)
    return loc.to_dict() if loc is not None else None


def register_resource(graph: Graph, locator: str, *, name: Optional[str] = None,
                      resource_id: Optional[str] = None,
                      url_type: Optional[str] = None,
                      description: Optional[str] = None) -> Dict[str, Any]:
    """Register a resource: assign a stable ID and store a locator (R0 STUB).

    Creates a ResourceNode carrying a new stable ID (a UUID unless ``resource_id`` is
    given) and the ``locator`` as its current ``url``, and adds it to the graph.
    This is the *identity + locator* half only — real ingest (copying bytes into
    an FS-index or MinIO store) is R1/R2, which will resolve the same ID through
    their backends. Returns ``{id, locator, kind}``."""
    import uuid
    from .nodes.resource_node import ResourceNode
    from .resources import classify_locator

    rid = resource_id or str(uuid.uuid4())
    node = ResourceNode(
        node_id=rid,
        name=name or "Unnamed Resource",
        url=locator or "",
        url_type=url_type or "",
        description=description or "",
    )
    graph.add_node(node)
    return {"id": rid, "locator": locator or "", "kind": classify_locator(locator or "")}


def scan_fs_resources(folder: str) -> List[Dict[str, Any]]:
    """Scan a folder with the FS-index backend (R1) and return its manifest
    entries as dicts (id, rel_path, name, resource_type, mtime, present).

    Pure/offline: files are indexed in place (Tropy-like) and minted stable IDs;
    it moves no bytes and touches no graph. Callers that want resolution build a
    registry and ``register`` the returned backend above passthrough."""
    from .resources import FSIndexBackend
    backend = FSIndexBackend(folder)
    backend.scan()
    return [e.to_dict() for e in backend.entries()]


def ingest_minio_resource(path: str, *, resource_id: Optional[str] = None,
                          config: Any = None) -> Dict[str, Any]:
    """Ingest a file into the SHARED MinIO/S3 store (R2) and return
    ``{id, object_key, s3_uri}``.

    ``resource_id`` (optional): reuse an EXISTING stable ID instead of minting a
    fresh one — this is how "promote to MinIO" keeps a resource's identity (the
    SAME id whether it lives on the file system or in MinIO — one ID space).

    By default the connection is read from the ``S3_*`` environment (the SAME
    vars Heriverse-Server uses → one shared object store); pass an explicit
    :class:`~s3dgraphy.resources.MinioConfig` to override. Needs the optional
    ``minio`` SDK (``pip install s3dgraphy[minio]``) AND a reachable server —
    raises :class:`MissingDependency` if the SDK is absent."""
    from .resources import MinioBackend, MinioConfig
    cfg = config or MinioConfig.from_env()
    backend = MinioBackend(cfg)
    rid, key = backend.ingest(path, resource_id=resource_id)
    return {"id": rid, "object_key": key, "s3_uri": f"s3://{cfg.bucket}/{key}"}


def presign_minio_resource(object_key: str, *, config: Any = None,
                           expires_seconds: int = 3600) -> Dict[str, Any]:
    """Presign a shared-MinIO ``object_key`` (as returned by
    :func:`ingest_minio_resource`) into a short-lived fetchable
    ``{object_key, http_url}``. Stateless (presign by key, no manifest). Connection
    from the ``S3_*`` env by default. Raises :class:`MissingDependency` without the
    ``minio`` SDK."""
    from .resources import MinioBackend, MinioConfig
    cfg = config or MinioConfig.from_env()
    loc = MinioBackend(cfg).presign_key(object_key, expires_seconds=expires_seconds)
    return {"object_key": object_key, "http_url": loc.value}


# ── DTC residency (R3: detach ↔ inject ↔ bake) ─────────────────────────────────
# A DTC can live WITH the asset store (a standalone record) rather than baked into
# em.json, and be baked back on demand. Resources are referenced by stable ID.
def detach_dtc(graph: Graph, process_id: str) -> Dict[str, Any]:
    """Extract the DTC anchored on ``process_id`` into a standalone JSON record
    (resources keyed by stable ID). Read-only — the graph is not mutated."""
    from .dtc import detach_dtc as _detach
    return _detach(graph, process_id)


def inject_dtc(graph: Graph, record: Dict[str, Any], *,
               injector_id: Optional[str] = None) -> Dict[str, Any]:
    """Re-create a DTC from a :func:`detach_dtc` record into ``graph``. Created
    nodes/edges are tagged ``injected_by`` (temporary); resources already present
    (by stable ID) are reused. Returns
    ``{injector_id, process_id, resource_ids, created}``."""
    from .dtc import inject_dtc as _inject
    return _inject(graph, record, injector_id=injector_id)


def bake_dtc(graph: Graph, injector_id: str) -> Dict[str, int]:
    """Promote an injected DTC to persistent (drop ``injected_by``) — the
    ``bake → em.json`` export. Returns ``{nodes, edges, overrides_cleared}``."""
    from .dtc import bake_dtc as _bake
    return _bake(graph, injector_id)


# ── Shelf substrate (Shelf v2 core: a collection of un-hatted resources) ───────
# A shelf-graph is a Graph of ResourceNode resources (R0 stable IDs), representable as
# a multigraph member OR a standalone reusable em.json. Each entry preserves its
# capability/origin for downstream tier badges. See :mod:`s3dgraphy.shelf`.
def new_shelf(graph_id: str = "shelf", name: Optional[str] = None) -> Graph:
    """Create an empty shelf-graph (tagged as a shelf collection)."""
    from .shelf import new_shelf as _n
    return _n(graph_id=graph_id, name=name)


def add_to_shelf(shelf: Graph, locator: str, *, resource_id: Optional[str] = None,
                 name: Optional[str] = None, url_type: Optional[str] = None,
                 description: Optional[str] = None,
                 resource_type: Optional[str] = None,
                 origin: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
    """Add a resource to the shelf (reuse-not-duplicate by ``resource_id``);
    ``origin`` = the capability/origin envelope, preserved. Returns the entry."""
    from .shelf import add_to_shelf as _a
    return _a(shelf, locator, resource_id=resource_id, name=name,
              url_type=url_type, description=description,
              resource_type=resource_type, origin=origin)


def list_shelf(shelf: Graph) -> List[Dict[str, Any]]:
    """List the shelf's resources — id/name/locator/kind + resource_type/origin
    (aligned with :func:`list_resources` / :func:`shelf_resources`)."""
    from .shelf import list_shelf as _l
    return _l(shelf)


def remove_from_shelf(shelf: Graph, resource_id: str) -> bool:
    """Remove a resource from the shelf. Returns True if it was present."""
    from .shelf import remove_from_shelf as _r
    return _r(shelf, resource_id)


def save_shelf(shelf: Graph, path: str) -> str:
    """Persist the shelf as a STANDALONE em.json file. Returns the path."""
    from .shelf import save_shelf as _s
    return _s(shelf, path)


def load_shelf(path: str) -> Tuple[Graph, List[str]]:
    """Load a standalone shelf em.json file → ``(graph, warnings)``."""
    from .shelf import load_shelf as _l
    return _l(path)


def instantiate_from_shelf(shelf: Graph, resource_id: str,
                           target_graph: Graph) -> Any:
    """Reference a shelf resource into ``target_graph`` by its stable ID
    (reuse-not-duplicate; capability/origin preserved). Returns the target node."""
    from .shelf import instantiate_from_shelf as _i
    return _i(shelf, resource_id, target_graph)


# Hatting facets. The ROLE picks the facet, and facets are NOT exclusive — the
# same Resource can be an RM of an epoch AND a Document in a paradata chain. Every
# facet keeps the P67 hinge (facet ─has_linked_resource→ ResourceNode); what differs is
# the edge towards what it represents / documents.
def hat_facets() -> Tuple[str, ...]:
    """The hatting facet names: ``("rm", "rmsf", "rmdoc", "document")``."""
    from .shelf import FACETS
    return FACETS


def attach_candidates(facet: str, graph: Graph) -> List[Dict[str, Any]]:
    """The nodes of ``graph`` a ``facet`` may attach to, as
    ``[{id, name, node_type, edge}]`` — derived from the datamodel's
    ``allowed_connections``, so a UI picker never hardcodes a type list. ``rm``
    yields the epochs in chronological order (first = ``has_first_epoch``)."""
    from .shelf import attach_candidates as _c
    return _c(facet, graph)
def hat_as_representation_model(target_graph: Graph, resource_id: str, *,
                                shelf: Optional[Graph] = None,
                                rm_id: Optional[str] = None,
                                name: Optional[str] = None,
                                epochs: Optional[List[str]] = None,
                                attach_to: Optional[str] = None) -> Dict[str, Any]:
    """Hat a shelf resource into ``target_graph`` as a RepresentationModel: the
    Resource is referenced by stable ID (R0 hinge) and an RM node references it via
    ``has_linked_resource`` (P67). An RM represents a STATE, so it binds to one or
    more **EpochNodes** (``epochs``, ordered: first → ``has_first_epoch``, rest →
    ``survive_in_epoch``); non-epoch targets are refused and returned in
    ``skipped``. ``attach_to`` is the deprecated single-epoch alias.
    Reuse-not-duplicate + idempotent. Returns
    ``{rm_id, resource_id, created, epochs, skipped, attached}``. No mesh import
    (EMTools does that)."""
    from .shelf import hat_as_representation_model as _h
    return _h(target_graph, resource_id, shelf=shelf, rm_id=rm_id, name=name,
              epochs=epochs, attach_to=attach_to)


def hat_as_rmsf(target_graph: Graph, resource_id: str, *,
                shelf: Optional[Graph] = None, rmsf_id: Optional[str] = None,
                name: Optional[str] = None,
                attach_to: Optional[str] = None) -> Dict[str, Any]:
    """Hat a shelf resource as a RepresentationModelSpecialFind: RMSF ─P67→ Resource
    plus ``SF ─has_representation_model_sf→ RMSF`` (P138i) when ``attach_to`` names a
    Special Find. Returns ``{rmsf_id, resource_id, created, attached}``."""
    from .shelf import hat_as_rmsf as _h
    return _h(target_graph, resource_id, shelf=shelf, rmsf_id=rmsf_id, name=name,
              attach_to=attach_to)


def hat_as_rmdoc(target_graph: Graph, resource_id: str, *,
                 shelf: Optional[Graph] = None, rmdoc_id: Optional[str] = None,
                 name: Optional[str] = None, attach_to: Optional[str] = None,
                 geometry: Optional[str] = None) -> Dict[str, Any]:
    """Hat a shelf resource as a RepresentationModelDoc — a Document instantiated in
    the 3D scene (e.g. a historical photo in place). RMDoc ─P67→ Resource plus
    ``Document ─has_representation_model_doc→ RMDoc`` (P138i) when ``attach_to``
    names a Document. No epoch, no stratigraphy: what grades an RMDoc is
    ``geometry``, the metric authority of its placement (Q-C) —
    ``reality_based → observable → asserted → symbolic``, ``em_based`` aside.
    Returns ``{rmdoc_id, resource_id, created, attached, geometry}``."""
    from .shelf import hat_as_rmdoc as _h
    return _h(target_graph, resource_id, shelf=shelf, rmdoc_id=rmdoc_id, name=name,
              attach_to=attach_to, geometry=geometry)


def hat_as_document(target_graph: Graph, resource_id: str, *,
                    shelf: Optional[Graph] = None, doc_id: Optional[str] = None,
                    name: Optional[str] = None, description: str = "",
                    role: Optional[str] = None,
                    content_nature: Optional[str] = None,
                    geometry: Optional[str] = None, mark_as_canonical: bool = True,
                    attach_to: Optional[str] = None) -> Dict[str, Any]:
    """Hat a shelf resource as a Document (E31) — a SOURCE, with no placement: the
    paradata entry point an ExtractorNode can later read from (``extracted_from``).
    Document ─P67→ Resource; ``attach_to`` picks its edge from the datamodel
    (Extractor → ``extracted_from``, stratigraphic → ``has_documentation``, other
    paradata → ``has_visual_reference``). ``doc_id`` naming an existing DocumentNode
    reuses it (one document shape with EMTools' ``create_master_document_node``).
    Returns ``{doc_id, resource_id, created, attached, attach_edge}``."""
    from .shelf import hat_as_document as _h
    return _h(target_graph, resource_id, shelf=shelf, doc_id=doc_id, name=name,
              description=description, role=role, content_nature=content_nature,
              geometry=geometry, mark_as_canonical=mark_as_canonical, attach_to=attach_to)


def hat_as_visual_resource(target_graph: Graph, resource_id: str, *,
                           shelf: Optional[Graph] = None,
                           attach_to: Optional[str] = None) -> Dict[str, Any]:
    """Reference a shelf resource as a VISUAL REFERENCE — the image a property is
    illustrated by. After BUGFIX-CONN2 ``has_visual_reference`` targets the
    resource-layer image (a ResourceNode, co-typed E36 Visual Item), not a source
    Document. No facet node: the visual resource IS the Resource, so this only
    references the ResourceNode and, when ``attach_to`` names a **PropertyNode**,
    wires ``PropertyNode ─has_visual_reference→ ResourceNode`` (P138i). A non-Property
    ``attach_to`` is refused (``attached=False``), never degraded. Returns
    ``{resource_id, created, attached, attach_edge}``."""
    from .shelf import hat_as_visual_resource as _h
    return _h(target_graph, resource_id, shelf=shelf, attach_to=attach_to)


def remove_shelf_resource(graph: Graph, resource_id: str) -> Dict[str, Any]:
    """Remove a shelf resource + clean up its now-orphan acquisition event (kept
    if the resource is still referenced, e.g. hatted). Returns
    ``{removed, referenced, events_removed}``."""
    from .shelf import remove_resource as _r
    return _r(graph, resource_id)


# ── Acquisition seam (Shelf v2 Session B: Tier-0 hook) ─────────────────────────
# An opaque source emits an AcquisitionDescriptor (versioned, canonical schema in
# JSON_config); s3Dgraphy consumes it into a Resource + a distinct acquisition DTC
# event (crmdig:D12) on the Shelf. Per-source JSON mappings (xlsx-import style)
# customize each repo's records. See :mod:`s3dgraphy.acquisition`.
def acquire_from_descriptor(descriptor: Any, shelf: Optional[Graph] = None
                            ) -> Tuple[Dict[str, Any], Graph]:
    """Tier-0 acquisition: descriptor (dict or AcquisitionDescriptor) → ``(info,
    shelf)``. Creates/reuses the Resource + its acquisition event on ``shelf``
    (a new shelf when ``None``). Raises if the descriptor carries a payload_graph
    (Tier 1/2 — later)."""
    from .acquisition import acquire_from_descriptor as _a
    return _a(descriptor, shelf)


def apply_acquisition_mapping(source_or_path: str,
                              record: Dict[str, Any]) -> Dict[str, Any]:
    """Translate a raw source ``record`` into an AcquisitionDescriptor dict using
    the per-source mapping (``source`` name or explicit path). Ercolano + fs ship."""
    from .acquisition import apply_mapping, load_mapping
    return apply_mapping(load_mapping(source_or_path), record)


def fs_acquisition_record(path: str) -> Dict[str, Any]:
    """Build a raw file-system record (filename/path/size/ext/media_type/…) for the
    ``fs`` mapping from a local ``path`` — the local project-folder acquisition source."""
    from .acquisition import fs_record
    return fs_record(path)


def acquisition_schema() -> Dict[str, Any]:
    """The canonical AcquisitionDescriptor JSON Schema (versioned)."""
    from .acquisition import schema
    return schema()


# ── connection resolution: REPORT-ONLY (S1) ───────────────────────────────────
# `Graph.validate_connection` is permissive by construction (it resolves the
# datamodel's CLASS names through the node_type-keyed map). These two ops measure
# what a CORRECT resolver would decide, changing nothing: no graph is mutated and
# `add_edge` keeps degrading exactly as it does today.
def connection_report(graph: Graph, *, max_cases: int = 0,
                      diagnose_generic_edges: bool = False) -> Dict[str, Any]:
    """What a correct resolver would decide about ``graph``'s edges. Returns
    ``{total_edges, resolved, would_degrade, already_generic, unknown_edge_type,
    dangling, delta, cases}`` — ``delta`` is the blast radius: edges that would
    degrade to ``generic_connection`` but are accepted by the current permissive
    core, i.e. exactly what a strict switch would change. With
    ``diagnose_generic_edges`` it also adds ``generic_diagnosis`` (see
    :func:`diagnose_generic_connections`). Read-only."""
    from .edges.connection_resolver import connection_report as _r
    return _r(graph, max_cases=max_cases,
              diagnose_generic_edges=diagnose_generic_edges)


def diagnose_generic_connections(graph: Graph, *, max_cases: int = 0) -> Dict[str, Any]:
    """Of the edges ALREADY typed ``generic_connection``, what type would their
    endpoints allow? Returns ``{total_generic, recoverable, ambiguous,
    no_candidate, dangling, cases}`` — ``recoverable`` = exactly one edge type
    fits, so the lost type is unambiguously reconstructible. **Diagnostic only**:
    nothing is re-typed and no graph is mutated."""
    from .edges.connection_resolver import diagnose_generic as _d
    return _d(graph, max_cases=max_cases)


def resolve_edge_type(source_node: Any, target_node: Any, declared_type: str) -> str:
    """The type an edge WOULD carry under correct resolution — the declared type
    when the datamodel allows it, else ``generic_connection``. Pure; does not
    touch any graph and does not change how edges are actually created."""
    from .edges.connection_resolver import resolve_edge_type as _r
    return _r(source_node, target_node, declared_type)


def _pick_narrative(graph: Graph, requested: Optional[str]) -> Optional[str]:
    """Which narrative a CLI command should act on, or ``None`` after explaining.

    One narrative needs no argument; several must be NAMED, and the error lists
    them rather than picking one — silently exporting the wrong text is the
    failure this refuses. Shared by every narrative command so they cannot answer
    the same question differently.
    """
    import sys

    if requested is not None:
        return requested
    narratives = [n for n in (getattr(graph, "nodes", []) or [])
                  if getattr(n, "node_type", None) == "narrative"]
    if len(narratives) == 1:
        return narratives[0].node_id
    if not narratives:
        print("error: this document contains no narrative", file=sys.stderr)
        return None
    print("error: several narratives — name one:", file=sys.stderr)
    for node in narratives:
        print(f"  {node.node_id}  {getattr(node, 'name', '')}", file=sys.stderr)
    return None


# ── thin CLI (part of the surface; no web deps) ────────────────────────────────
def main(argv: Optional[List[str]] = None) -> int:
    """`python -m s3dgraphy.api <op> ...` — a thin CLI over the ops above."""
    import argparse
    import json
    import sys

    # The exception classes, fetched through the PACKAGE path rather than used as
    # locals. Run as `python -m s3dgraphy.api`, this module is `__main__`, and a
    # submodule that does `from ..api import MissingDependency` imports a SECOND
    # copy of it under its real name: two distinct class objects for one name, so
    # `except MissingDependency` here would not catch what the exporter raised —
    # a traceback instead of the intended exit code. Importing it the same way the
    # raiser does makes the two identities the same object under both entry points.
    from .api import MissingDependency as _MissingDependency

    ap = argparse.ArgumentParser(prog="s3dgraphy.api", description=__doc__.splitlines()[0])
    sub = ap.add_subparsers(dest="op", required=True)
    sub.add_parser("open", help="parse an .em.json file, print stats").add_argument("path")
    sub.add_parser("validate", help="validate an .em.json file").add_argument("path")
    sub.add_parser("project-ttl", help=".em.json → Turtle (stdout)").add_argument("path")
    sub.add_parser("graphml", help=".em.json → GraphML (stdout)").add_argument("path")
    sub.add_parser("import-graphml", help="GraphML → .em.json (stdout)").add_argument("path")
    cv = sub.add_parser("convert",
                        help="GraphML → .em.json FILE + conversion report "
                             "(the 1.7 migration path; deterministic)")
    cv.add_argument("path", help="input .graphml")
    cv.add_argument("-o", "--output", default=None,
                    help="output .em.json (default: alongside the input, "
                         "same stem; '-' writes the document to stdout)")
    cv.add_argument("--json", action="store_true",
                    help="machine-readable report on stdout")
    cv.add_argument("--list-warnings", action="store_true",
                    help="print the individual importer warnings, not just counts")
    cv.add_argument("--max-warnings", type=int, default=20,
                    help="truncate the warning list (0 = all; default 20)")
    cv.add_argument("--force", action="store_true",
                    help="overwrite the output file if it already exists")
    nl = sub.add_parser("export-narrative-latex",
                        help="narrative → .tex + .bib (sources become bib entries)")
    nl.add_argument("path", help="input .em.json")
    nl.add_argument("narrative_id", nargs="?", default=None,
                    help="which narrative (default: the only one, or list them)")
    nl.add_argument("-o", "--output", default=None,
                    help="output stem: writes <stem>.tex and <stem>.bib "
                         "(default: alongside the input; '-' writes the tex to "
                         "stdout and the bib to stderr, so the two can be piped "
                         "apart)")
    nl.add_argument("--force", action="store_true",
                    help="overwrite the output files if they already exist")
    nd = sub.add_parser("export-narrative-docx",
                        help="narrative → .docx (needs the [docx] extra)")
    nd.add_argument("path", help="input .em.json")
    nd.add_argument("narrative_id", nargs="?", default=None,
                    help="which narrative (default: the only one, or list them)")
    # A FLAG and not a second positional, like export-narrative-latex: with two
    # optional positionals, `... doc.em.json out.docx` silently reads the output
    # path as the narrative id and fails with a baffling "no narrative node with
    # id '/path/out.docx'".
    nd.add_argument("-o", "--output", default=None,
                    help="output .docx (default: alongside the input)")
    nd.add_argument("--force", action="store_true",
                    help="overwrite the output file if it already exists")
    r = sub.add_parser("resolve", help="resolve an authority term")
    r.add_argument("term")
    r.add_argument("facet")
    sub.add_parser("list-resources", help="list a graph's resources (LinkNodes)").add_argument("path")
    rr = sub.add_parser("resolve-resource", help="resolve a resource ID → Location")
    rr.add_argument("path")
    rr.add_argument("resource_id")
    sub.add_parser("scan-resources", help="FS-index scan a folder → manifest").add_argument("folder")
    cr = sub.add_parser("connection-report",
                        help="REPORT-ONLY: what a correct connection resolver "
                             "would decide about an .em.json (changes nothing)")
    cr.add_argument("path")
    cr.add_argument("--json", action="store_true", help="machine-readable output")
    cr.add_argument("--max-cases", type=int, default=0,
                    help="truncate the case list (0 = all)")
    cr.add_argument("--diagnose-generic", action="store_true",
                    help="also ask what type the already-generic edges would "
                         "take, judging by their endpoints (diagnostic only)")
    dd = sub.add_parser("detach-dtc", help="extract a DTC → standalone JSON record")
    dd.add_argument("path")
    dd.add_argument("process_id")
    # ── Shelf substrate (all operate on a standalone shelf em.json file) ──────
    sn = sub.add_parser("shelf-new", help="create an empty standalone shelf em.json")
    sn.add_argument("path")
    sn.add_argument("--graph-id", default="shelf")
    sn.add_argument("--name", default=None)
    sub.add_parser("shelf-list", help="list a shelf's resources").add_argument("path")
    sa = sub.add_parser("shelf-add", help="add a resource to a shelf (saved back)")
    sa.add_argument("path")
    sa.add_argument("locator")
    sa.add_argument("--resource-id", default=None)
    sa.add_argument("--name", default=None)
    sa.add_argument("--resource-type", default=None)
    sa.add_argument("--origin-repo", default=None, help="source repo id (origin)")
    sa.add_argument("--capabilities", default=None,
                    help="comma-separated source capabilities, e.g. genesis,interpretation")
    sa.add_argument("--scope", default=None, help="payload scope (origin)")
    srm = sub.add_parser("shelf-remove", help="remove a resource from a shelf (saved back)")
    srm.add_argument("path")
    srm.add_argument("resource_id")
    si = sub.add_parser("shelf-instantiate",
                        help="reference a shelf resource into a target em.json (saved back)")
    si.add_argument("path")
    si.add_argument("resource_id")
    si.add_argument("target")
    # ── Acquisition (Tier 0): descriptor / mapped record → Resource + event on a shelf ──
    aq = sub.add_parser("acquire", help="Tier-0 acquire a descriptor onto a shelf")
    aq.add_argument("descriptor", help="path to an AcquisitionDescriptor .json")
    aq.add_argument("--shelf", required=True, help="shelf em.json (created if missing)")
    am = sub.add_parser("acquire-map",
                        help="map a raw source record → descriptor, then acquire onto a shelf")
    am.add_argument("source", help="per-source mapping name (e.g. ercolano) or path")
    am.add_argument("record", help="path to a raw source record .json")
    am.add_argument("--shelf", required=True, help="shelf em.json (created if missing)")
    args = ap.parse_args(argv)

    if args.op in ("open", "validate", "project-ttl", "graphml"):
        with open(args.path, encoding="utf-8") as f:
            doc = json.load(f)
        graph, warnings = load_emjson(doc)
        for w in warnings:
            print(f"warning: {w}", file=sys.stderr)
        if args.op == "open":
            print(json.dumps(validate(graph)["stats"]))
        elif args.op == "validate":
            print(json.dumps(validate(graph), indent=2))
        elif args.op == "project-ttl":
            print(project_ttl(graph))
        elif args.op == "graphml":
            print(graph_to_graphml(graph))
    elif args.op == "export-narrative-latex":
        with open(args.path, encoding="utf-8") as f:
            doc = json.load(f)
        graph, warnings = load_emjson(doc)
        for w in warnings:
            print(f"warning: {w}", file=sys.stderr)
        target_id = _pick_narrative(graph, args.narrative_id)
        if target_id is None:
            return 1
        try:
            out = export_narrative_latex(graph, target_id)
        except KeyError as exc:
            print(f"error: {exc}", file=sys.stderr)
            return 1
        if args.output == "-":
            # tex on stdout, bib on stderr: two artefacts, one process, pipeable
            # apart — the same convention `convert` uses for doc/report.
            print(out["tex"], end="")
            print(out["bib"], end="", file=sys.stderr)
        else:
            stem = Path(args.output) if args.output else Path(args.path).with_suffix("")
            tex_path = stem.with_suffix(".tex")
            bib_path = stem.with_suffix(".bib")
            existing = [p for p in (tex_path, bib_path) if p.exists()]
            if existing and not args.force:
                for p in existing:
                    print(f"error: {p} exists (use --force)", file=sys.stderr)
                return 1
            tex_path.write_text(out["tex"], encoding="utf-8")
            bib_path.write_text(out["bib"], encoding="utf-8")
            print(f"{tex_path}\n{bib_path}")
    elif args.op == "export-narrative-docx":
        with open(args.path, encoding="utf-8") as f:
            doc = json.load(f)
        graph, warnings = load_emjson(doc)
        for w in warnings:
            print(f"warning: {w}", file=sys.stderr)
        target_id = _pick_narrative(graph, args.narrative_id)
        if target_id is None:
            return 1
        out_path = (Path(args.output) if args.output
                    else Path(args.path).with_suffix(".docx"))
        if out_path.exists() and not args.force:
            print(f"error: {out_path} exists (use --force)", file=sys.stderr)
            return 1
        # Baked ONCE and rendered from that: baking again for the warning list
        # would double the file reads and leave two snapshots free to disagree
        # about what was resolved. Relative image locators are relative to the
        # em.json, not to the directory the command ran from.
        try:
            baked = bake_narrative(graph, target_id,
                                   base_dir=str(Path(args.path).parent))
        except KeyError as exc:
            print(f"error: {exc}", file=sys.stderr)
            return 1
        # Said before the write, so it is visible even when the render then fails.
        for ref in baked.unresolved:
            print(f"warning: unresolved embed {ref}", file=sys.stderr)
        try:
            from .exporter.docx_exporter import render_docx
            blob = render_docx(baked)
        except _MissingDependency as exc:
            # Distinct exit code: "this build cannot" is not the same failure as
            # "your input is wrong", and a script driving the CLI should be able
            # to tell them apart without parsing the message.
            print(f"error: {exc}", file=sys.stderr)
            return 3
        out_path.write_bytes(blob)
        print(out_path)
    elif args.op == "import-graphml":
        print(json.dumps(graphml_to_emjson(Path(args.path).read_bytes())))
    elif args.op == "convert":
        src = Path(args.path)
        doc, report = convert_graphml_to_emjson(src.read_bytes())
        if args.output == "-":
            # Document on stdout, report on stderr, so the two can be piped
            # apart. Same serialisation as the file branch — a conversion must
            # be byte-identical whichever way it leaves the process.
            print(json.dumps(doc, ensure_ascii=False, indent=2, sort_keys=True))
            out_path = None
        else:
            out_path = Path(args.output) if args.output else src.with_suffix(".em.json")
            if out_path.exists() and not args.force:
                print(f"error: {out_path} already exists (use --force to overwrite)",
                      file=sys.stderr)
                return 1
            out_path.parent.mkdir(parents=True, exist_ok=True)
            with open(out_path, "w", encoding="utf-8") as f:
                json.dump(doc, f, ensure_ascii=False, indent=2, sort_keys=True)
                f.write("\n")
        stream = sys.stderr if args.output == "-" else sys.stdout
        if args.json:
            print(json.dumps({**report,
                              "input": str(src),
                              "output": str(out_path) if out_path else None},
                             indent=2), file=stream)
        else:
            if out_path:
                print(f"wrote {out_path}", file=stream)
            print(format_conversion_report(report,
                                           max_warnings=args.max_warnings,
                                           list_warnings=args.list_warnings),
                  file=stream)
    elif args.op == "resolve":
        print(json.dumps(resolve_authority(args.term, args.facet), indent=2))
    elif args.op in ("list-resources", "resolve-resource"):
        with open(args.path, encoding="utf-8") as f:
            doc = json.load(f)
        graph, warnings = load_emjson(doc)
        for w in warnings:
            print(f"warning: {w}", file=sys.stderr)
        if args.op == "list-resources":
            print(json.dumps(list_resources(graph), indent=2))
        else:
            print(json.dumps(resolve_resource(graph, args.resource_id), indent=2))
    elif args.op == "scan-resources":
        print(json.dumps(scan_fs_resources(args.folder), indent=2))
    elif args.op == "connection-report":
        with open(args.path, encoding="utf-8") as f:
            doc = json.load(f)
        graph, warnings = load_emjson(doc)
        for w in warnings:
            print(f"warning: {w}", file=sys.stderr)
        rep = connection_report(graph, max_cases=args.max_cases,
                                diagnose_generic_edges=args.diagnose_generic)
        if args.json:
            print(json.dumps(rep, indent=2))
        else:
            from .edges.connection_resolver import format_connection_report
            print(format_connection_report(rep))
    elif args.op == "detach-dtc":
        with open(args.path, encoding="utf-8") as f:
            doc = json.load(f)
        graph, warnings = load_emjson(doc)
        for w in warnings:
            print(f"warning: {w}", file=sys.stderr)
        print(json.dumps(detach_dtc(graph, args.process_id), indent=2))
    elif args.op == "shelf-new":
        shelf = new_shelf(graph_id=args.graph_id, name=args.name)
        save_shelf(shelf, args.path)
        print(json.dumps({"ok": True, "path": args.path, "graph_id": args.graph_id}))
    elif args.op == "shelf-list":
        shelf, warnings = load_shelf(args.path)
        for w in warnings:
            print(f"warning: {w}", file=sys.stderr)
        print(json.dumps(list_shelf(shelf), indent=2))
    elif args.op == "shelf-add":
        shelf, warnings = load_shelf(args.path)
        for w in warnings:
            print(f"warning: {w}", file=sys.stderr)
        origin = None
        if args.origin_repo or args.capabilities or args.scope:
            origin = {}
            if args.origin_repo:
                origin["repo"] = args.origin_repo
            if args.capabilities:
                origin["capabilities"] = [c.strip() for c in args.capabilities.split(",") if c.strip()]
            if args.scope:
                origin["scope"] = args.scope
        entry = add_to_shelf(shelf, args.locator, resource_id=args.resource_id,
                             name=args.name, resource_type=args.resource_type,
                             origin=origin)
        save_shelf(shelf, args.path)
        print(json.dumps(entry, indent=2))
    elif args.op == "shelf-remove":
        shelf, warnings = load_shelf(args.path)
        for w in warnings:
            print(f"warning: {w}", file=sys.stderr)
        removed = remove_from_shelf(shelf, args.resource_id)
        save_shelf(shelf, args.path)
        print(json.dumps({"ok": True, "removed": removed}))
    elif args.op == "shelf-instantiate":
        shelf, sw = load_shelf(args.path)
        for w in sw:
            print(f"warning: {w}", file=sys.stderr)
        target, tw = load_emjson_file(args.target)
        for w in tw:
            print(f"warning: {w}", file=sys.stderr)
        node = instantiate_from_shelf(shelf, args.resource_id, target)
        from .exporter.emjson_exporter import export_emjson
        export_emjson(target, args.target)
        print(json.dumps({"ok": True, "resource_id": node.node_id,
                          "target": args.target}))
    elif args.op in ("acquire", "acquire-map"):
        import os as _os
        if args.op == "acquire":
            with open(args.descriptor, encoding="utf-8") as f:
                descriptor = json.load(f)
        else:  # acquire-map
            with open(args.record, encoding="utf-8") as f:
                record = json.load(f)
            descriptor = apply_acquisition_mapping(args.source, record)
        # load-or-create the shelf
        if _os.path.isfile(args.shelf):
            shelf, sw = load_shelf(args.shelf)
            for w in sw:
                print(f"warning: {w}", file=sys.stderr)
        else:
            shelf = new_shelf()
        info, shelf = acquire_from_descriptor(descriptor, shelf)
        save_shelf(shelf, args.shelf)
        out = {"ok": True, "shelf": args.shelf, **info}
        if args.op == "acquire-map":
            out["descriptor"] = descriptor
        print(json.dumps(out, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
