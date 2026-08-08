"""TRAD1 · multilingual datamodel descriptions/labels — the mechanism.

The node datamodel (`s3Dgraphy_node_datamodel.json`) carries English-only
`description`/`label` per class. This keeps a **sidecar** translations file next
to it (`datamodel_translations.json`) so the English stays the git-diffable
source and translations are added without restructuring the datamodel (its many
consumers keep reading `description` as a string; `--check` stays green).

Shape (keyed by CLASS — the stable datamodel identity that EMStudio's
`typeDescription`/node_type→class lookup already uses)::

    {
      "schema": "s3Dgraphy_datamodel_translations",
      "version": "1.0",
      "languages": ["en", "it", ...],
      "entries": {
        "<Class>": {
          "description": {"en": "...", "it": "...", "validated_it": false},
          "label":       {"en": "...", ...}
        }
      }
    }

en is the SOURCE (seeded from the datamodel, never edited via xlsx). Other
languages are translation surface; each carries a per-key ``validated_<lang>``
flag (coverage ≠ validation, like the UI i18n). xlsx is only an editing surface
for native speakers, round-tripped by this script.

CLI::

    python -m s3dgraphy.tools.datamodel_i18n seed         # (re)seed en from the datamodel
    python -m s3dgraphy.tools.datamodel_i18n export a.xlsx
    python -m s3dgraphy.tools.datamodel_i18n import a.xlsx
    python -m s3dgraphy.tools.datamodel_i18n --check      # en in sync with the datamodel?
"""

from __future__ import annotations

import json
import sys
from pathlib import Path
from typing import Any, Dict, List, Tuple

_HERE = Path(__file__).resolve().parent
_JSON_CONFIG = _HERE.parent / "JSON_config"
DATAMODEL = _JSON_CONFIG / "s3Dgraphy_node_datamodel.json"
TRANSLATIONS = _JSON_CONFIG / "datamodel_translations.json"

# Languages mirror the EMStudio UI locales (en = source/default).
LANGUAGES: List[str] = ["en", "it", "el", "he", "es", "pl", "ro", "fr"]
FIELDS = ("description", "label")


def _load(path: Path) -> Dict[str, Any]:
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def _collect_en(datamodel: Dict[str, Any]) -> Dict[str, Dict[str, str]]:
    """Walk the datamodel and return {class: {field: english_text}} for every
    class entry that declares a description or a label (recursing into
    ``subtypes``). The class is the entry's ``class`` value."""
    out: Dict[str, Dict[str, str]] = {}

    def walk(node: Any) -> None:
        if isinstance(node, dict):
            cls = node.get("class")
            if isinstance(cls, str):
                fields = {}
                for f in FIELDS:
                    v = node.get(f)
                    if isinstance(v, str) and v.strip():
                        fields[f] = v
                if fields:
                    out.setdefault(cls, {}).update(fields)
            for v in node.values():
                walk(v)
        elif isinstance(node, list):
            for v in node:
                walk(v)

    walk(datamodel)
    return out


def seed(write: bool = True) -> Dict[str, Any]:
    """(Re)generate the sidecar: refresh every `en` from the datamodel while
    PRESERVING existing translations and their validated flags. New classes/
    fields are added; nothing non-en is dropped."""
    en = _collect_en(_load(DATAMODEL))
    existing = _load(TRANSLATIONS) if TRANSLATIONS.exists() else {}
    entries: Dict[str, Any] = existing.get("entries", {}) if isinstance(existing, dict) else {}

    for cls in sorted(en):
        cls_entry = entries.setdefault(cls, {})
        for field, en_text in en[cls].items():
            fe = cls_entry.setdefault(field, {})
            fe["en"] = en_text  # en is always the datamodel's (source of truth)
    doc = {
        "schema": "s3Dgraphy_datamodel_translations",
        "version": str(existing.get("version", "1.0")) if isinstance(existing, dict) else "1.0",
        "languages": LANGUAGES,
        "entries": entries,
    }
    if write:
        with open(TRANSLATIONS, "w", encoding="utf-8") as f:
            json.dump(doc, f, ensure_ascii=False, indent=1)
            f.write("\n")
    return doc


def check() -> int:
    """Exit 0 iff every datamodel en is present and identical in the sidecar."""
    if not TRANSLATIONS.exists():
        print("datamodel_translations.json missing — run: seed")
        return 1
    en = _collect_en(_load(DATAMODEL))
    doc = _load(TRANSLATIONS)
    entries = doc.get("entries", {})
    drift = []
    for cls, fields in en.items():
        for field, text in fields.items():
            got = entries.get(cls, {}).get(field, {}).get("en")
            if got != text:
                drift.append(f"{cls}.{field}")
    if drift:
        print(f"en drift ({len(drift)}): {', '.join(drift[:8])}"
              f"{' …' if len(drift) > 8 else ''} — run: seed")
        return 1
    print(f"datamodel translations en in sync ({len(en)} classes).")
    return 0


def _rows(doc: Dict[str, Any]) -> List[Tuple[str, str]]:
    """(class, field) keys present, stable order."""
    keys: List[Tuple[str, str]] = []
    for cls in sorted(doc.get("entries", {})):
        for field in FIELDS:
            if field in doc["entries"][cls]:
                keys.append((cls, field))
    return keys


def export_xlsx(path: str) -> None:
    """One row per (class, field): a column per language + validated_<lang>."""
    from openpyxl import Workbook

    doc = _load(TRANSLATIONS)
    langs = doc.get("languages", LANGUAGES)
    wb = Workbook()
    ws = wb.active
    ws.title = "translations"
    header = ["class", "field"]
    for lang in langs:
        header.append(lang)
        if lang != "en":
            header.append(f"validated_{lang}")
    ws.append(header)
    for cls, field in _rows(doc):
        fe = doc["entries"][cls][field]
        row = [cls, field]
        for lang in langs:
            row.append(fe.get(lang, ""))
            if lang != "en":
                row.append(bool(fe.get(f"validated_{lang}", False)))
        ws.append(row)
    wb.save(path)
    print(f"exported {len(_rows(doc))} keys × {len(langs)} langs → {path}")


def import_xlsx(path: str, write: bool = True) -> Dict[str, Any]:
    """Read the xlsx back into the sidecar. NEVER overwrites `en` (the source);
    a differing `en` cell is reported as a conflict and ignored. Sets the
    translations + validated flags. Idempotent."""
    from openpyxl import load_workbook

    doc = _load(TRANSLATIONS)
    entries = doc["entries"]
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["translations"] if "translations" in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        print("empty xlsx")
        return doc
    header = [str(h) if h is not None else "" for h in rows[0]]
    idx = {name: i for i, name in enumerate(header)}
    conflicts, updated = [], 0
    for r in rows[1:]:
        cls = r[idx["class"]] if idx.get("class") is not None else None
        field = r[idx["field"]] if idx.get("field") is not None else None
        if not cls or not field:
            continue
        fe = entries.get(str(cls), {}).get(str(field))
        if fe is None:
            continue  # unknown key — the datamodel drives which keys exist
        for lang in doc.get("languages", LANGUAGES):
            if lang not in idx:
                continue
            val = r[idx[lang]]
            val = "" if val is None else str(val)
            if lang == "en":
                if val and val != fe.get("en"):
                    conflicts.append(f"{cls}.{field}")  # en is source: don't touch
                continue
            if val:
                if fe.get(lang) != val:
                    updated += 1
                fe[lang] = val
            vkey = f"validated_{lang}"
            if vkey in idx:
                fe[vkey] = str(r[idx[vkey]]).strip().lower() in ("1", "true", "yes", "x")
    if write:
        with open(TRANSLATIONS, "w", encoding="utf-8") as f:
            json.dump(doc, f, ensure_ascii=False, indent=1)
            f.write("\n")
    print(f"imported {path}: {updated} translation cell(s) updated"
          + (f"; {len(conflicts)} en-conflict(s) ignored: {conflicts[:5]}" if conflicts else ""))
    return doc


def main(argv: List[str]) -> int:
    if not argv or argv[0] in ("-h", "--help"):
        print(__doc__)
        return 0
    cmd = argv[0]
    if cmd == "--check":
        return check()
    if cmd == "seed":
        seed()
        print(f"seeded en for {len(_load(TRANSLATIONS)['entries'])} classes → {TRANSLATIONS.name}")
        return 0
    if cmd == "export":
        export_xlsx(argv[1] if len(argv) > 1 else "datamodel_translations.xlsx")
        return 0
    if cmd == "import":
        if len(argv) < 2:
            print("usage: import <path.xlsx>")
            return 2
        import_xlsx(argv[1])
        return 0
    print(f"unknown command {cmd!r}")
    return 2


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
