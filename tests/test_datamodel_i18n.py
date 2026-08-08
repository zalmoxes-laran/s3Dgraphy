"""TRAD1 — multilingual datamodel descriptions (sidecar + xlsx round-trip).

The English source lives in the node datamodel; a sidecar
`datamodel_translations.json` (keyed by CLASS) adds other languages with a
per-key `validated_<lang>` flag. This tests the mechanism: en is seeded from the
datamodel and never lost, xlsx export→import round-trips translations, and en is
never overwritten from the xlsx.
"""

import json
import shutil

import pytest

from s3dgraphy.tools import datamodel_i18n as t


def test_collect_en_covers_the_datamodel_classes():
    en = t._collect_en(t._load(t.DATAMODEL))
    assert len(en) >= 40
    su = en.get("StratigraphicUnit", {})
    assert su.get("description") and su.get("label")


def test_committed_sidecar_en_is_in_sync():
    """The shipped sidecar's en must match the datamodel (── the `--check`)."""
    assert t.check() == 0


def test_round_trip_and_en_not_overwritten(tmp_path, monkeypatch):
    # work on a temp copy of the real sidecar so the committed file is untouched
    tr = tmp_path / "datamodel_translations.json"
    shutil.copyfile(t.TRANSLATIONS, tr)
    monkeypatch.setattr(t, "TRANSLATIONS", tr)

    # preset an it translation + validated flag, plus a bogus en to try to sneak in
    doc = json.loads(tr.read_text())
    doc["entries"]["StratigraphicUnit"]["description"]["it"] = "Unità stratigrafica."
    doc["entries"]["StratigraphicUnit"]["description"]["validated_it"] = True
    source_en = doc["entries"]["StratigraphicUnit"]["description"]["en"]
    tr.write_text(json.dumps(doc, ensure_ascii=False))

    xlsx = tmp_path / "t.xlsx"
    t.export_xlsx(str(xlsx))

    # tamper the xlsx en cell for that key → import must IGNORE it (en is source)
    from openpyxl import load_workbook
    wb = load_workbook(str(xlsx))
    ws = wb["translations"]
    hdr = [c.value for c in ws[1]]
    ci = {h: i for i, h in enumerate(hdr)}
    for row in ws.iter_rows(min_row=2):
        if row[ci["class"]].value == "StratigraphicUnit" and row[ci["field"]].value == "description":
            row[ci["en"]].value = "HACKED EN"
    wb.save(str(xlsx))

    t.import_xlsx(str(xlsx))

    out = json.loads(tr.read_text())["entries"]["StratigraphicUnit"]["description"]
    assert out["en"] == source_en          # en never overwritten from xlsx
    assert out["it"] == "Unità stratigrafica."   # it round-tripped
    assert out["validated_it"] is True           # validated flag round-tripped


def test_seed_is_idempotent_and_preserves_translations(tmp_path, monkeypatch):
    tr = tmp_path / "datamodel_translations.json"
    shutil.copyfile(t.TRANSLATIONS, tr)
    monkeypatch.setattr(t, "TRANSLATIONS", tr)
    doc = json.loads(tr.read_text())
    doc["entries"]["StratigraphicUnit"]["description"]["it"] = "Preservami."
    tr.write_text(json.dumps(doc, ensure_ascii=False))

    t.seed()  # re-seed en

    out = json.loads(tr.read_text())["entries"]["StratigraphicUnit"]["description"]
    assert out["it"] == "Preservami."   # non-en survives a re-seed
    assert t.check() == 0
