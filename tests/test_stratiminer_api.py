"""StratiMiner's two halves, and the seam between them.

The pipeline is::

    folder of documents  --(AI)-->  em_data.xlsx  --(deterministic)-->  em.json

These tests are about the properties that make that split worth having, not
about the importer's row handling (``test_lossless_roundtrip.py`` covers that).
Each one names a way the arrangement can quietly stop being true:

* the second arrow claims to be **deterministic** — if it is not, the table
  stops being a reviewable intermediate: you cannot diff two generations of it,
  and anything keyed on ``node_id`` detaches on every re-import while the graph
  still *looks* correct;
* the derived ids must be scoped per graph — otherwise two sites that both
  number a unit ``U1`` mint the same id and a merge silently fuses them;
* the AI half must be asked for the **table**, never for the graph. That is a
  property of a prompt, i.e. of a string, and a string is exactly the kind of
  thing that gets edited by someone who does not know why it said what it said.

No binary fixture is shipped: the workbook is written with openpyxl (a hard
dependency of s3dgraphy) so the fixture is readable in the diff.
"""

from __future__ import annotations

import json
import os
import sys

import pytest

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "src"))

from s3dgraphy import api  # noqa: E402


_HEADERS = {
    "Units": ("ID", "TYPE", "NAME"),
    "Epochs": ("ID", "NAME", "START", "END", "COLOR"),
    "Claims": ("TARGET_ID", "PROPERTY_TYPE", "VALUE", "TARGET2_ID",
               "UNITS", "EXTRACTOR_1", "DOCUMENT_1", "AUTHOR_1"),
    "Authors": ("ID", "KIND", "DISPLAY_NAME", "ORCID", "AFFILIATION"),
    "Documents": ("ID", "FILENAME", "TITLE", "YEAR", "AUTHOR_IDS",
                  "ROLE", "CONTENT_NATURE", "GEOMETRY"),
}

# One of each thing the table can carry: two units and a virtual one, an epoch,
# a scalar quale with an attribution chain, a SECOND quale of the same type on
# the same unit (two readings — see test_repeated_property_type_...), and a
# physical relation.
_FIXTURE = {
    "Units": [
        ("U1", "US", "Wall West"),
        ("U2", "US", "Wall East"),
        ("U3", "USVs", "Reconstructed Frieze"),
    ],
    "Epochs": [
        ("E1", "Roman", -50, 100, "#aa5500"),
    ],
    "Claims": [
        ("U1", "has_first_epoch", "E1", "", "", "", "", ""),
        ("U2", "has_first_epoch", "E1", "", "", "", "", ""),
        ("U1", "length", "14.5", "", "m", "Measured on site", "D.01", "A.01"),
        ("U1", "length", "15.0", "", "m", "Second reading", "D.01", "A.01"),
        ("U1", "overlies", "", "U2", "", "", "", ""),
    ],
    "Authors": [
        ("A.01", "author", "Jane Roe", "", "ISPC-CNR"),
    ],
    "Documents": [
        ("D.01", "report.pdf", "Excavation Report", "2024", "A.01",
         "analytical", "2d_object", "observable"),
    ],
}


def _write_workbook(path, fixture=None, sheets=None):
    """Materialise a fixture dict as a real .xlsx."""
    from openpyxl import Workbook

    fixture = _FIXTURE if fixture is None else fixture
    wb = Workbook()
    wb.remove(wb.active)
    for name in (sheets if sheets is not None else _HEADERS):
        ws = wb.create_sheet(name)
        ws.append(_HEADERS[name])
        for row in fixture.get(name, []):
            ws.append(row)
    wb.save(str(path))
    return str(path)


@pytest.fixture
def em_data(tmp_path):
    return _write_workbook(tmp_path / "em_data.xlsx")


# ── the deterministic half ────────────────────────────────────────────────────

def test_import_em_data_is_reproducible(em_data):
    """Same workbook in, same document out — byte for byte.

    This is the property that lets ``em_data.xlsx`` be a reviewable
    intermediate. Before the ids were derived from the sheet keys, the importer
    minted a fresh ``uuid4`` per node, so re-importing an unchanged table
    produced a document that differed everywhere while asserting exactly the
    same archaeology: nothing failed, and every id-keyed thing (a saved layout,
    an annotation, a from-sketch arrangement) came loose on each run.
    """
    first = api.import_em_data(em_data, graph_id="probe")
    second = api.import_em_data(em_data, graph_id="probe")
    assert json.dumps(first, sort_keys=True) == json.dumps(second,
                                                           sort_keys=True)


def test_graph_id_scopes_the_derived_ids(em_data):
    """The same unit number in two graphs must not become the same node.

    Derived ids buy reproducibility, but keyed on the sheet value ALONE they
    would make ``U1`` of one site and ``U1`` of another collide — and a merge
    would fuse two distinct walls without a word. ``graph_id`` is in the key to
    prevent exactly that.
    """
    here = api.import_em_data(em_data, graph_id="site-a")
    there = api.import_em_data(em_data, graph_id="site-b")
    ids_here = {n["id"] for n in here["graph"]["nodes"]}
    ids_there = {n["id"] for n in there["graph"]["nodes"]}
    assert ids_here.isdisjoint(ids_there)


def test_repeated_property_type_is_not_overwritten(em_data):
    """Two ``length`` claims on the same unit stay two nodes.

    A unit legitimately carries two readings of the same measurement. Keying the
    derived id on target+property_type alone would mint ONE id for both, and the
    second row would land on top of the first — a row silently swallowed, which
    is worse than a row rejected.
    """
    graph, warnings, _stats = api.em_data_to_graph(em_data, graph_id="probe")
    lengths = [n for n in graph.nodes
               if getattr(n, "property_type", None) == "length"]
    assert len(lengths) == 2, warnings
    assert len({n.node_id for n in lengths}) == 2
    assert {n.value for n in lengths} == {"14.5", "15.0"}


def test_stats_report_what_was_read(em_data):
    """``stats`` is why the table is inspectable: it says what came IN, per
    sheet, next to what came out. Derived from the finished graph instead, a
    skipped row and a row that never existed look identical."""
    _graph, _warnings, stats = api.em_data_to_graph(em_data, graph_id="probe")
    assert stats["rows_units"] == len(_FIXTURE["Units"])
    assert stats["rows_epochs"] == len(_FIXTURE["Epochs"])
    assert stats["rows_claims"] == len(_FIXTURE["Claims"])
    assert stats["rows_authors"] == len(_FIXTURE["Authors"])
    assert stats["rows_documents"] == len(_FIXTURE["Documents"])
    assert stats["nodes_total"] > 0 and stats["edges_total"] > 0


def test_missing_sheet_names_the_sheet(tmp_path):
    """A four-sheet workbook must fail at the door, saying which sheet is
    absent. Failing later, inside a parser, sends the user looking at their
    data when the problem is the file's shape."""
    path = _write_workbook(
        tmp_path / "partial.xlsx",
        sheets=("Units", "Epochs", "Claims", "Authors"))
    with pytest.raises(Exception) as exc:
        api.em_data_to_graph(path, graph_id="partial")
    assert "Documents" in str(exc.value)


def test_em_data_sheets_matches_the_importer():
    """The five names a caller enumerates must come from the importer, not from
    a second list that is free to drift from it."""
    from s3dgraphy.importer.unified_xlsx_importer import UnifiedXLSXImporter
    assert api.em_data_sheets() == tuple(UnifiedXLSXImporter._SHEETS)


# ── the AI half: what the model is asked for ──────────────────────────────────

def test_prompt_asks_for_the_table_never_the_graph():
    """The invariant of the whole arrangement, asserted against the prompt.

    The AI produces ``em_data.xlsx`` and nothing else. If someone ever edits the
    template to ask for em.json or a GraphML directly — which would look like a
    helpful shortcut — the reviewable step disappears and an unvalidated guess
    lands straight in the language's own format, where a wrong node type is
    indistinguishable from a right one. Nothing else in the suite would notice.
    """
    prompt = api.stratiminer_prompt()
    assert "em_data.xlsx" in prompt
    lowered = prompt.lower()
    assert ".graphml" not in lowered
    assert "em.json" not in lowered


def test_prompt_names_all_five_sheets():
    """The prompt has to describe the shape the deterministic half expects; a
    sheet added to the importer and forgotten in the template would be asked of
    nobody and arrive empty."""
    prompt = api.stratiminer_prompt()
    for sheet in api.em_data_sheets():
        assert sheet in prompt, f"sheet {sheet} not described in the prompt"


def test_prompt_carries_the_documents_folder(tmp_path):
    """When a folder is given, the model has to be told where the sources are —
    otherwise it invents filenames, and invented provenance is the one error
    this pipeline exists to prevent."""
    folder = tmp_path / "DosCo"
    folder.mkdir()
    prompt = api.stratiminer_prompt(documents_folder=str(folder))
    assert str(folder) in prompt


def test_prompt_is_pure():
    """Building a prompt calls no model and needs no key: it is a string
    builder. Two calls with the same options give the same string."""
    assert api.stratiminer_prompt(language="it") == api.stratiminer_prompt(
        language="it")


def test_every_declared_column_is_described_in_the_prompt():
    """The code's header layout and the template's tables must agree.

    Two descriptions of one shape: ``_COLUMNS`` (what a writer emits and the
    importer reads) and the markdown tables in the prompt (what the model is
    told to produce). Nothing links them, so a column added to one and not the
    other would be asked of nobody, or written and never mentioned — and the
    workbook would look complete either way.
    """
    prompt = api.stratiminer_prompt()
    for sheet, header in api.em_data_columns().items():
        for column in header:
            assert f"`{column}`" in prompt, (
                f"{sheet}.{column} is declared in _COLUMNS but not described "
                f"in StratiMiner_Extraction_Prompt.md")


# ── the materialiser: rows → em_data.xlsx ─────────────────────────────────────

def test_write_em_data_round_trips_through_the_importer(tmp_path):
    """Rows written by the materialiser must come back out of the importer.

    This is the join between StratiMiner's two arrows: what Path A writes is
    what the deterministic half reads. If the writer's header order or naming
    drifted from the importer's expectations the workbook would still open, and
    every row would import as empty.
    """
    out = tmp_path / "written.xlsx"
    report = api.write_em_data({
        "Units": [{"ID": "U1", "TYPE": "US", "NAME": "Wall West"},
                  {"ID": "U2", "TYPE": "US", "NAME": "Wall East"}],
        "Epochs": [{"ID": "E1", "NAME": "Roman", "START": -50, "END": 100}],
        "Claims": [{"TARGET_ID": "U1", "PROPERTY_TYPE": "has_first_epoch",
                    "VALUE": "E1"},
                   {"TARGET_ID": "U1", "PROPERTY_TYPE": "overlies",
                    "TARGET2_ID": "U2"}],
        "Authors": [],
        "Documents": [],
    }, str(out))

    assert report["rows"]["Units"] == 2
    assert report["warnings"] == []

    graph, warnings, stats = api.em_data_to_graph(str(out), graph_id="written")
    assert stats["rows_units"] == 2
    names = {n.name for n in graph.nodes}
    assert {"U1", "U2"} <= names, warnings
    assert any(e.edge_type == "overlies" for e in graph.edges)


def test_write_em_data_drops_invented_columns_loudly(tmp_path):
    """A column the model invented is dropped WITH a warning.

    Silently carrying it into the workbook would be the worst option: the
    importer ignores unknown columns, so the value would look captured and be
    nowhere. The table exists to be checked, and a warning is what makes the
    check possible.
    """
    out = tmp_path / "invented.xlsx"
    report = api.write_em_data({
        "Units": [{"ID": "U1", "TYPE": "US", "CONFIDENCE": "high"}],
    }, str(out))

    assert any("CONFIDENCE" in w for w in report["warnings"])
    from openpyxl import load_workbook
    header = [c.value for c in next(load_workbook(str(out))["Units"].rows)]
    assert "CONFIDENCE" not in header
    assert header == list(api.em_data_columns()["Units"])


def test_write_em_data_always_writes_all_five_sheets(tmp_path):
    """Even with nothing to put in them: the importer fails fast on a missing
    sheet, so a materialiser that omitted the empty ones would produce a file
    that cannot be imported at all."""
    out = tmp_path / "sparse.xlsx"
    api.write_em_data({"Units": [{"ID": "U1", "TYPE": "US"}]}, str(out))
    from openpyxl import load_workbook
    assert set(load_workbook(str(out)).sheetnames) == set(api.em_data_sheets())


def test_write_em_data_joins_list_values(tmp_path):
    """A model returning ``["A.01","A.02"]`` for a comma-separated column gets
    joined, not rejected: openpyxl cannot store a list, and the importer already
    reads that column as a comma-separated string."""
    out = tmp_path / "lists.xlsx"
    api.write_em_data({
        "Documents": [{"ID": "D.01", "FILENAME": "r.pdf",
                       "AUTHOR_IDS": ["A.01", "A.02"]}],
    }, str(out))
    from openpyxl import load_workbook
    ws = load_workbook(str(out))["Documents"]
    header = [c.value for c in next(ws.rows)]
    row = [c.value for c in list(ws.rows)[1]]
    assert row[header.index("AUTHOR_IDS")] == "A.01, A.02"
